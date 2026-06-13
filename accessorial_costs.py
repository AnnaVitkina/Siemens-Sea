"""Build and format the Accessorial Costs tab for rate card output."""

from __future__ import annotations

from pathlib import Path

import pandas as pd

from carrier_lookup import (
    build_all_carriers_apply_if,
    build_carrier_name_apply_if,
    carrier_code_from_filename,
    detect_carrier_key,
    get_biofuel_cost_names,
)
from config import (
    ACCESSORIAL_COSTS_SHEET_NAME,
    ADDON_SMF_TAB,
    FINANCING_ACCESSORIAL_COST_NAME,
    GLOSSARY_TAB,
    LCL_TMP_ACCESSORIAL_COST_NAME,
    TMP_ACCESSORIAL_COST_NAME,
)
from extractor import SubfolderSelection, extract_sheet_to_dataframe
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from glossary_lookup import (
    GlossaryFeeLookup,
    _glossary_fee_text,
    load_glossary_fee_lookup,
    parse_financing_fee,
    parse_lcl_tmp_fee,
    parse_tmp_consol_fee,
    parse_tmp_fees,
)

ACCESSORIAL_COLUMNS = [
    "Cost Name",
    "Currency",
    "Price",
    "Apply if",
    "Rate by",
]

TMP_RATE_BY = "Quantity/Container"
BIOFUEL_RATE_BY = "ACC/TEU"
BIOFUEL_COST_TYPE_PATTERN = "biofuel"

TABLE_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")
TABLE_THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def apply_table_sheet_formatting(
    worksheet,
    *,
    header_row: int,
    column_count: int,
    data_end_row: int,
    column_max_widths: dict[int, int] | None = None,
    default_max_width: int = 40,
    min_column_width: int = 10,
) -> None:
    header_font = Font(bold=True, size=10)
    body_font = Font(size=10)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for row_index in range(1, data_end_row + 1):
        for column_index in range(1, column_count + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.border = TABLE_THIN_BORDER
            if row_index == header_row:
                cell.font = header_font
                cell.fill = TABLE_HEADER_FILL
                cell.alignment = header_alignment
            else:
                cell.font = body_font
                cell.alignment = body_alignment

    for column_index in range(1, column_count + 1):
        column_letter = get_column_letter(column_index)
        max_length = 0
        for row_index in range(1, data_end_row + 1):
            value = worksheet.cell(row=row_index, column=column_index).value
            if value is None:
                continue
            max_length = max(max_length, len(str(value).split("\n")[0]))

        configured_max = (column_max_widths or {}).get(column_index, default_max_width)
        worksheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, min_column_width),
            configured_max,
        )

    worksheet.freeze_panes = worksheet.cell(row=header_row + 1, column=1)


def _format_price_number(value: float) -> str:
    if float(value).is_integer():
        return str(int(value))
    return str(value)


def _format_tmp_price(unit_rate: float, max_cap: float | None) -> str:
    unit_text = _format_price_number(unit_rate)
    if max_cap is None:
        return f"{unit_text} p/unit"
    return f"{unit_text} p/unit, {_format_price_number(max_cap)} MAX"


def _last_three_chars(value: object) -> str | None:
    if pd.isna(value):
        return None
    code = str(value).strip().upper()
    return code[-3:] if len(code) >= 3 else code or None


def _build_biofuel_apply_if(
    row: pd.Series,
    carrier_code: str,
) -> str:
    parts = [
        f"Origin Country equals {row['Origin Country']}",
        f"Origin Post equals {_last_three_chars(row['Origin Location'])}",
        f"Destination country equals {row['Destination Country']}",
        f"Destination post equals {_last_three_chars(row['Destination Location'])}",
        f"Carrier Name equals {carrier_code}",
    ]
    return ", and ".join(parts)


def _glossary_rows_from_lookup(
    lookup: GlossaryFeeLookup,
    shipper: str,
    source_file: Path,
    flow: str,
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    carrier_code = lookup.carrier_name
    glossary_df = extract_sheet_to_dataframe(source_file, GLOSSARY_TAB)
    tmp_text = _glossary_fee_text(glossary_df, "TMP-Fee-ALL_IN")
    financing_text = _glossary_fee_text(glossary_df, "Financing Fee")

    if tmp_text:
        for entry in parse_tmp_fees(tmp_text, shipper):
            rows.append(
                {
                    "Cost Name": TMP_ACCESSORIAL_COST_NAME,
                    "Currency": entry.currency,
                    "Price": _format_tmp_price(entry.unit_rate, entry.max_cap),
                    "Apply if": _carrier_apply_if(
                        carrier_code,
                        shipper,
                        flow,
                        entry.apply_if_suffix,
                    ),
                    "Rate by": TMP_RATE_BY,
                }
            )
        consol_entry = parse_tmp_consol_fee(tmp_text)
        if consol_entry is not None:
            rows.append(
                {
                    "Cost Name": TMP_ACCESSORIAL_COST_NAME,
                    "Currency": consol_entry.currency,
                    "Price": _format_tmp_price(consol_entry.unit_rate, consol_entry.max_cap),
                    "Apply if": _carrier_apply_if(carrier_code, shipper, flow),
                    "Rate by": TMP_RATE_BY,
                }
            )

    if financing_text:
        financing = parse_financing_fee(financing_text)
        if financing is not None:
            rows.extend(
                [
                    {
                        "Cost Name": f"{FINANCING_ACCESSORIAL_COST_NAME} (20' Container)",
                        "Currency": financing.currency,
                        "Price": _format_price_number(financing.rate_20),
                        "Apply if": _carrier_apply_if(carrier_code, shipper, flow),
                        "Rate by": "Container/20_FT",
                    },
                    {
                        "Cost Name": f"{FINANCING_ACCESSORIAL_COST_NAME} (40'/45' Container)",
                        "Currency": financing.currency,
                        "Price": _format_price_number(financing.rate_40),
                        "Apply if": _carrier_apply_if(carrier_code, shipper, flow),
                        "Rate by": "Container/40_FT",
                    },
                ]
            )

    return rows


def _carrier_apply_if(
    carrier_code: str,
    shipper: str,
    flow: str,
    suffix: str = "",
) -> str:
    if flow == "BCN":
        base = build_all_carriers_apply_if(shipper, flow=flow)
        if suffix:
            return f"{base}; {suffix}"
        return base
    return build_carrier_name_apply_if(carrier_code, suffix)


def _biofuel_cost_name(
    carrier_key: str,
    cost_type: object,
    shipper: str,
    flow: str,
) -> str:
    biofuel_names = get_biofuel_cost_names(shipper, flow=flow)
    if carrier_key in biofuel_names:
        return biofuel_names[carrier_key]
    if pd.notna(cost_type):
        return str(cost_type).strip()
    return "Add-on Biofuel"


def _biofuel_rows_from_file(
    source_file: Path,
    carrier_code: str,
    carrier_key: str,
    shipper: str,
    flow: str,
) -> list[dict[str, str]]:
    addon_df = extract_sheet_to_dataframe(source_file, ADDON_SMF_TAB)
    if "Cost Type" not in addon_df.columns:
        return []

    biofuel_df = addon_df[
        addon_df["Cost Type"]
        .astype(str)
        .str.contains(BIOFUEL_COST_TYPE_PATTERN, case=False, na=False)
    ]
    if biofuel_df.empty:
        return []

    price_column = None
    for candidate in ("Preis per TEU", "Price for reduction of 1 ton CO2e"):
        if candidate in biofuel_df.columns:
            price_column = candidate
            break
    if price_column is None:
        return []

    currency_column = "Currentcy" if "Currentcy" in biofuel_df.columns else "Currency"
    rows: list[dict[str, str]] = []
    for _, row in biofuel_df.iterrows():
        price_value = row.get(price_column)
        if pd.isna(price_value):
            continue

        currency_value = row.get(currency_column) if currency_column in biofuel_df.columns else None
        if pd.isna(currency_value):
            currency_value = "EUR"

        rows.append(
            {
                "Cost Name": _biofuel_cost_name(
                    carrier_key,
                    row.get("Cost Type"),
                    shipper,
                    flow,
                ),
                "Currency": str(currency_value).strip().upper(),
                "Price": _format_price_number(float(price_value)),
                "Apply if": (
                    build_all_carriers_apply_if(shipper, flow=flow)
                    if flow == "BCN"
                    else _build_biofuel_apply_if(row, carrier_code)
                ),
                "Rate by": BIOFUEL_RATE_BY,
            }
        )
    return rows


def _lcl_glossary_rows_from_file(
    source_file: Path,
    carrier_code: str,
) -> list[dict[str, str]]:
    glossary_df = extract_sheet_to_dataframe(source_file, GLOSSARY_TAB)
    tmp_text = _glossary_fee_text(glossary_df, "TMP-Fee-ALL_IN")
    entry = parse_lcl_tmp_fee(tmp_text)
    if entry is None:
        return []

    return [
        {
            "Cost Name": LCL_TMP_ACCESSORIAL_COST_NAME,
            "Currency": entry.currency,
            "Price": _format_tmp_price(entry.unit_rate, entry.max_cap),
            "Apply if": build_carrier_name_apply_if(carrier_code),
            "Rate by": TMP_RATE_BY,
        }
    ]


def build_lcl_accessorial_costs_dataframe(
    shipper: str,
    individual_selections: list[SubfolderSelection],
) -> pd.DataFrame:
    rows: list[dict[str, str]] = []

    for selection in individual_selections:
        if GLOSSARY_TAB not in selection.tabs:
            continue

        carrier_code = carrier_code_from_filename(
            selection.file_path.name,
            shipper=shipper,
            flow="FCL",
        )
        carrier_key = detect_carrier_key(
            selection.file_path.name,
            shipper=shipper,
            flow="FCL",
        )
        if not carrier_code or not carrier_key:
            continue

        rows.extend(_lcl_glossary_rows_from_file(selection.file_path, carrier_code))

    if not rows:
        return pd.DataFrame(columns=ACCESSORIAL_COLUMNS)

    return pd.DataFrame(rows, columns=ACCESSORIAL_COLUMNS)


def write_accessorial_costs_sheet(workbook: Workbook, accessorial_df: pd.DataFrame) -> None:
    worksheet = workbook.create_sheet(title=ACCESSORIAL_COSTS_SHEET_NAME)
    for column_index, column_name in enumerate(ACCESSORIAL_COLUMNS, start=1):
        worksheet.cell(row=1, column=column_index, value=column_name)

    data_rows = list(dataframe_to_rows(accessorial_df, index=False, header=False))
    for row_offset, row in enumerate(data_rows, start=2):
        for column_offset, value in enumerate(row, start=1):
            worksheet.cell(row=row_offset, column=column_offset, value=value)

    data_end_row = max(1, 1 + len(data_rows))
    apply_table_sheet_formatting(
        worksheet,
        header_row=1,
        column_count=len(ACCESSORIAL_COLUMNS),
        data_end_row=data_end_row,
        column_max_widths={
            1: 48,
            3: 24,
            4: 56,
            5: 22,
        },
    )


def build_accessorial_costs_dataframe(
    shipper: str,
    individual_selections: list[SubfolderSelection],
    flow: str = "FCL",
) -> pd.DataFrame:
    rows: list[dict[str, str]] = []

    for selection in individual_selections:
        carrier_code = carrier_code_from_filename(
            selection.file_path.name,
            shipper=shipper,
            flow=flow,
        )
        carrier_key = detect_carrier_key(
            selection.file_path.name,
            shipper=shipper,
            flow=flow,
        )
        if not carrier_code or not carrier_key:
            continue

        if GLOSSARY_TAB in selection.tabs:
            lookup = load_glossary_fee_lookup(
                shipper,
                source_file=selection.file_path,
                flow=flow,
            )
            if lookup is not None:
                rows.extend(
                    _glossary_rows_from_lookup(lookup, shipper, selection.file_path, flow)
                )

        if ADDON_SMF_TAB in selection.tabs and selection.file_path.exists():
            sheet_names = pd.ExcelFile(selection.file_path).sheet_names
            if ADDON_SMF_TAB in sheet_names:
                rows.extend(
                    _biofuel_rows_from_file(
                        selection.file_path,
                        carrier_code,
                        carrier_key,
                        shipper,
                        flow,
                    )
                )

    if not rows:
        return pd.DataFrame(columns=ACCESSORIAL_COLUMNS)

    return pd.DataFrame(rows, columns=ACCESSORIAL_COLUMNS)
