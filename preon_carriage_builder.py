"""Build Pre/on carriage (per carrier) rate cards."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from accessorial_costs import apply_table_sheet_formatting
from carrier_lookup import carrier_code_from_filename, detect_carrier_key
from config import ACCESSORIAL_COSTS_SHEET_NAME, GLOSSARY_TAB, OUTPUT_DIR, RATE_CARD_SHEET_NAME
from extractor import SubfolderSelection, extract_sheet_to_dataframe, slugify
from glossary_lookup import _glossary_fee_text
from thc_lookup import FclThcLookup, load_fcl_thc_lookup

PREON_TAB = "Pre-On-Carriage_RoW"
PREON_GENERIC_TAB = "PreOnCarriage_Containerized_EU"
PREON_DIESELFLOATER_TAB = "Emergency Dieselfloater Pre_On"
PREON_GENERIC_SERVICES_TAB = "PreOn_Containerized_EU_Services"
PREON_GENERIC_ADD_SERVICES_TABS = (
    "Add_Services_Glomb_Br. Hafenb",
    "Add_Services_Glomb_Br. Hafenb.",
)
PREON_GENERIC_TERMS_TAB_SHORT = {
    "HAPAG_Terms & Conditions": "HLCU",
    "MAERSK_Terms & Condition": "MAEU",
    "MSC_Terms & Conditions": "MSCU",
    "ONE_Terms & Conditions": "ONEY",
}

SHIPMENT_COLUMNS = [
    "From Country",
    "From (UN/LOCODE)",
    "From(Location/Port)",
    "Origin city",
    "To Country",
    "To (UN/LOCODE)",
    "to (Location/Port)",
    "Destination city",
    "SERVICE",
    "Valid From",
    "Valid To",
]

BOLD_SHIPMENT_HEADERS = {
    "From Country",
    "From (UN/LOCODE)",
    "Origin city",
    "To Country",
    "To (UN/LOCODE)",
    "to (Location/Port)",
    "Destination city",
}

GENERIC_BOLD_SHIPMENT_HEADERS: set[str] = set()

PREON_ACCESSORIAL_COLUMNS = [
    "Cost Name",
    "Rate by",
    "Apply over",
    "Cost",
    "Valid From",
    "Valid To",
]
HEALTHINEERS_SPECIAL_ACCESSORIAL_PREFIXES = (
    "Empty Container Fee",
    "Pre-Carriage",
    "Pickup Fee",
    "Drop Off Fee",
)


@dataclass(frozen=True)
class CostGroup:
    key: str
    cost_name: str
    apply_if: str
    rate_by: str
    rate_types: tuple[str, ...]
    include_min: bool
    include_max: bool
    sort_priority: int = 500
    highlight_red: bool = False


TYPE_MODE_ORDER = {"FCL": 0, "FTL": 1, "LTL": 2}
SERVICE_ORDER = {"ONCARRIAGE": 0, "PRE-CARRIAGE": 1}


def _normalize_number(value: object) -> object:
    if pd.isna(value):
        return None
    if isinstance(value, (int, float)):
        if float(value).is_integer():
            return int(value)
        return float(value)
    text = str(value).strip().replace(",", ".")
    try:
        number = float(text)
    except ValueError:
        return value
    if float(number).is_integer():
        return int(number)
    return number


def _number_format(value: object) -> str:
    if isinstance(value, (int, float)) and float(value).is_integer():
        return "0"
    if isinstance(value, float):
        return "0.##########"
    return "General"


def _format_number_text(value: float) -> str:
    if float(value).is_integer():
        return str(int(value))
    return str(value)


def _format_date(value: object) -> str | None:
    if pd.isna(value):
        return None
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        text = str(value).strip()
        return text or None
    return parsed.strftime("%d.%m.%Y")


def _last_three(value: object) -> str | None:
    if pd.isna(value):
        return None
    code = str(value).strip().upper()
    return code[-3:] if len(code) >= 3 else code or None


def _value_from_aliases(row: pd.Series, aliases: tuple[str, ...]) -> object:
    for alias in aliases:
        if alias in row.index:
            return row.get(alias)
    return None


def _equipment_apply_if(max_weight: object, min_weight_excluded: object | None = None) -> str:
    parts = ["Equipment Type equals 'FTL/STANDARD'"]
    if max_weight is not None:
        parts.append(f"Weight/kg less than or equal to '{_normalize_number(max_weight)}'")
    if min_weight_excluded is not None:
        parts.append(f"Weight/kg greater than {_normalize_number(min_weight_excluded)}")
    return "; and ".join(parts)


def _round_weight_lower_bound(value: object | None) -> object | None:
    number = _normalize_number(value)
    if not isinstance(number, (int, float)):
        return number
    if number <= 0:
        return None
    return int(number // 1000 * 1000)


def _submode_to_rateby(sub_mode: object) -> str:
    text = "UNKNOWN" if pd.isna(sub_mode) else _normalize_sub_mode(sub_mode)
    return f"Container/{text}"


def _normalize_sub_mode(sub_mode: object) -> str:
    text = str(sub_mode).strip().upper().replace("GP", "FT")
    # Normalize equivalent container notations, e.g. 20` FT, 20_FT, 20FT -> 20FT.
    return "".join(ch for ch in text if ch.isalnum())


def _is_container_like_sub_mode(sub_mode_normalized: str) -> bool:
    if not sub_mode_normalized:
        return False
    if sub_mode_normalized in {"FTL", "LTL"}:
        return False
    return any(ch.isdigit() for ch in sub_mode_normalized)


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    text = str(value).strip()
    if text.lower() in {"", "nan", "none"}:
        return ""
    return text


def _ltl_apply_if(max_weight: object, min_weight_excluded: object | None = None) -> str:
    parts = ["Equipment Type equals 'LTL/STANDARD'"]
    if max_weight is not None:
        parts.append(f"Weight/kg less than or equal to '{_normalize_number(max_weight)}'")
    if min_weight_excluded is not None:
        parts.append(f"Weight/kg greater than '{_normalize_number(min_weight_excluded)}'")
    return "; and ".join(parts)


def _weight_bounds_apply_if(max_weight: object, min_weight_excluded: object | None = None) -> str:
    parts: list[str] = []
    if max_weight is not None:
        parts.append(f"Weight/kg less than or equal to '{_normalize_number(max_weight)}'")
    if min_weight_excluded is not None:
        parts.append(f"Weight/kg greater than '{_normalize_number(min_weight_excluded)}'")
    return "; and ".join(parts)


def _default_unmapped_apply_if(
    type_mode: str,
    max_weight: object,
    min_weight_excluded: object | None,
) -> str:
    bounds = _weight_bounds_apply_if(max_weight, min_weight_excluded)
    if type_mode == "FTL":
        base = "Equipment Type equals 'FTL/STANDARD'"
        return f"{base}; and {bounds}" if bounds else base
    if type_mode == "LTL":
        base = "Equipment Type equals 'LTL/STANDARD'"
        return f"{base}; and {bounds}" if bounds else base
    return bounds


def _build_cost_groups(row: pd.Series) -> list[CostGroup]:
    type_mode = _normalize_text(row.get("Type Mode", "")).upper()
    calc_basis_type = _normalize_text(row.get("Calculation Basis Type", "")).lower()
    cost_type = _normalize_text(row.get("Cost Type"))
    sub_mode = row.get("Sub Mode")
    max_weight = _normalize_number(row.get("Maximum weight (included)"))
    min_weight = _normalize_number(row.get("Minimum weight (excluded)"))
    min_cost = _normalize_number(row.get("Min Cost Value"))
    max_cost = _normalize_number(row.get("Max Cost Value"))
    include_min = min_cost is not None
    include_max = max_cost is not None
    sub_text = _normalize_text(sub_mode) or "N/A"
    sub_mode_normalized = _normalize_text(sub_mode).upper()
    max_weight_text = "N/A" if max_weight is None else str(max_weight)
    min_weight_rounded = _round_weight_lower_bound(min_weight)
    min_weight_text = "N/A" if min_weight_rounded is None else str(min_weight_rounded)
    groups: list[CostGroup] = []

    if type_mode == "FCL":
        normalized_sub_mode = _normalize_sub_mode(sub_mode)
        groups.append(
            CostGroup(
                key=f"FCL|{normalized_sub_mode}|container",
                cost_name="Transport cost (FCL truck)",
                apply_if="",
                rate_by=_submode_to_rateby(sub_mode),
                rate_types=("p/unit",),
                include_min=False,
                include_max=False,
                sort_priority=0,
            )
        )
        return groups

    if cost_type in {"Delta Surcharge", "Energy Surcharge"}:
        normalized_sub_mode = _normalize_sub_mode(sub_mode) if sub_mode_normalized else "UNKNOWN"
        groups.append(
            CostGroup(
                key=f"SURCHARGE|{cost_type.upper()}|{normalized_sub_mode}",
                cost_name=f"{cost_type} ({normalized_sub_mode})",
                apply_if="",
                rate_by=f"Container/{normalized_sub_mode}",
                rate_types=("p/unit",),
                include_min=include_min,
                include_max=include_max,
                sort_priority=250,
            )
        )
        return groups

    # Handle rows where Type Mode is blank but Sub Mode is a container code (40_GP, reefer, etc).
    # These should follow FCL-style container transport cost instead of UNMAPPED.
    normalized_sub_mode = _normalize_sub_mode(sub_mode) if sub_mode_normalized else ""
    if (
        (type_mode == "" or type_mode.startswith("FCL"))
        and calc_basis_type in {"truck", "container"}
        and _is_container_like_sub_mode(normalized_sub_mode)
    ):
        groups.append(
            CostGroup(
                key=f"FCL|{normalized_sub_mode}|container",
                cost_name="Transport cost (FCL truck)",
                apply_if="",
                rate_by=f"Container/{normalized_sub_mode}",
                rate_types=("p/unit",),
                include_min=False,
                include_max=False,
                sort_priority=0,
            )
        )
        return groups

    # Explicit FTL truck patterns requested by user.
    is_ftl_truck_special = (
        calc_basis_type == "truck"
        and max_weight is not None
        and (
            (type_mode == "" and sub_mode_normalized in {"FTL", "LTL"})
            or (
                type_mode == "FTL"
                and sub_mode_normalized in {"", "FTL", "10 TO AND ABOVE"}
            )
        )
    )
    if is_ftl_truck_special:
        groups.append(
            CostGroup(
                key=f"FTL|TRUCK_SPECIAL|{max_weight}|{min_weight_rounded}",
                cost_name=f"Transport cost (FTL 10 truck <= {max_weight_text} (min - {min_weight_text}))",
                apply_if=_equipment_apply_if(max_weight, min_weight_rounded),
                rate_by="per shipment",
                rate_types=("Flat",),
                include_min=False,
                include_max=False,
                sort_priority=10,
            )
        )
        return groups

    # Explicit FTL per ton (Type Mode empty, Sub Mode FTL, min/max cost present).
    if (
        type_mode == ""
        and sub_mode_normalized == "FTL"
        and include_min
        and include_max
    ):
        groups.append(
            CostGroup(
                key=f"FTL|PER_TON_SPECIAL|{max_weight}|{min_weight_rounded}",
                cost_name=f"Transport cost (FTL per ton <= {max_weight_text}(min - {min_weight_text}))",
                apply_if=_equipment_apply_if(max_weight, min_weight_rounded),
                rate_by="Weight/kg",
                rate_types=("p/1000 unit",),
                include_min=True,
                include_max=True,
                sort_priority=20,
            )
        )
        return groups

    # Keep existing explicit FTL truck naming for regular FTL truck rows.
    if type_mode == "FTL" and calc_basis_type == "truck":
        if sub_mode_normalized == "10 TO AND ABOVE" and max_weight is None:
            return groups
        if max_weight is None:
            return groups
        sub_text = "FTL" if pd.isna(sub_mode) or not str(sub_mode).strip() else str(sub_mode).strip()
        groups.append(
            CostGroup(
                key=f"FTL|DEFAULT|{sub_text}|{max_weight}|{min_weight_rounded}|{include_min}|{include_max}",
                cost_name=f"Transport cost (FTL {sub_text} {max_weight_text})",
                apply_if=_equipment_apply_if(max_weight, min_weight_rounded),
                rate_by="per shipment",
                rate_types=("Flat",),
                include_min=include_min,
                include_max=include_max,
                sort_priority=30,
            )
        )
        return groups

    is_ltl_context = type_mode == "LTL" or (type_mode == "" and sub_mode_normalized == "LTL")
    if is_ltl_context:
        if calc_basis_type == "shipment":
            groups.append(
                CostGroup(
                    key=f"LTL|SHIPMENT|{max_weight}|{min_weight_rounded}",
                    cost_name=f"Transport cost (LTL per shipment <= {max_weight_text})",
                    apply_if="",
                    rate_by="per shipment",
                    rate_types=("p/unit",),
                    include_min=include_min,
                    include_max=include_max,
                    sort_priority=100,
                )
            )
            return groups

        if calc_basis_type == "kg":
            if include_min or include_max:
                groups.append(
                    CostGroup(
                        key=f"LTL|KG|WITH_MINMAX|{max_weight}|{min_weight_rounded}",
                        cost_name="Transport cost (LTL per kg)",
                        apply_if="",
                        rate_by="Weight/kg",
                        rate_types=("p/unit",),
                        include_min=include_min,
                        include_max=include_max,
                        sort_priority=120,
                    )
                )
            else:
                groups.append(
                    CostGroup(
                        key=f"LTL|KG|NO_MINMAX|{max_weight}|{min_weight_rounded}",
                        cost_name="Transport cost (LTL per kg (min/max=0)",
                        apply_if="",
                        rate_by="Weight/kg",
                        rate_types=("p/unit",),
                        include_min=False,
                        include_max=False,
                        sort_priority=121,
                    )
                )
            return groups

        if calc_basis_type == "w/m":
            if include_min and include_max and min_cost == max_cost:
                groups.append(
                    CostGroup(
                        key="LTL|WM|SHIPMENT_FLAT",
                        cost_name="Transport cost (LTL per w/m (min=max))",
                        apply_if="",
                        rate_by="per shipment",
                        rate_types=("Flat",),
                        include_min=False,
                        include_max=False,
                        sort_priority=131,
                    )
                )
            else:
                groups.append(
                    CostGroup(
                        key="LTL|WM|VOLUME",
                        cost_name="Transport cost (LTL per w/m)",
                        apply_if="",
                        rate_by="Volume/cbm",
                        rate_types=("p/unit",),
                        include_min=include_min,
                        include_max=include_max,
                        sort_priority=130,
                    )
                )
            return groups

        if calc_basis_type == "ton":
            groups.append(
                CostGroup(
                    key=f"LTL|TON|{max_weight}|{min_weight_rounded}",
                    cost_name=f"Transport cost (LTL per ton <= {max_weight_text})",
                    apply_if=_ltl_apply_if(max_weight, min_weight_rounded),
                    rate_by="Weight/kg",
                    rate_types=("p/1000 unit",),
                    include_min=include_min,
                    include_max=include_max,
                    sort_priority=110,
                )
            )
            return groups

        if calc_basis_type == "lbs":
            groups.append(
                CostGroup(
                    key=f"LTL|LBS|{max_weight}|{min_weight_rounded}",
                    cost_name=f"Transport cost (LTL per ton <= {max_weight_text})",
                    apply_if=_ltl_apply_if(max_weight, min_weight_rounded),
                    rate_by="Weight/kg",
                    rate_types=("p/unit",),
                    include_min=include_min,
                    include_max=include_max,
                    sort_priority=140,
                )
            )
            return groups

    # Generic ton behavior for non-LTL/non-FTL rows that still have ton basis.
    if calc_basis_type == "ton":
        # Normalize empty Type Mode + Sub Mode LTL into the main LTL per-ton column.
        if is_ltl_context:
            groups.append(
                CostGroup(
                    key=f"LTL|TON|{max_weight}|{min_weight_rounded}",
                    cost_name=f"Transport cost (LTL per ton <= {max_weight_text})",
                    apply_if=_ltl_apply_if(max_weight, min_weight_rounded),
                    rate_by="Weight/kg",
                    rate_types=("p/1000 unit",),
                    include_min=include_min,
                    include_max=include_max,
                    sort_priority=110,
                )
            )
            return groups

        groups.append(
            CostGroup(
                key=(
                    f"TON|{type_mode or 'EMPTY'}|{sub_text}|{max_weight}|{min_weight_rounded}|"
                    f"{include_min}|{include_max}"
                ),
                cost_name=f"Transport cost ({sub_text} per ton {max_weight_text})",
                apply_if=_equipment_apply_if(max_weight, min_weight_rounded),
                rate_by="Weight/kg",
                rate_types=("p/1000 unit",),
                include_min=include_min,
                include_max=include_max,
                sort_priority=200,
            )
        )
        return groups

    # Fallback for all unmapped combinations: append at end and highlight in red.
    groups.append(
        CostGroup(
            key=(
                f"UNMAPPED|{type_mode or 'EMPTY'}|{sub_mode_normalized or 'EMPTY'}|"
                f"{calc_basis_type or 'EMPTY'}|{max_weight}|{include_min}|{include_max}"
            ),
            cost_name=(
                f"Transport cost (UNMAPPED {type_mode or 'EMPTY'} / "
                f"{sub_mode_normalized or 'EMPTY'} / {calc_basis_type or 'EMPTY'})"
            ),
            apply_if=_default_unmapped_apply_if(type_mode, max_weight, min_weight_rounded),
            rate_by="per shipment" if calc_basis_type in {"truck", "shipment"} else "Weight/kg",
            rate_types=("Flat" if calc_basis_type in {"truck", "shipment"} else "p/unit",),
            include_min=include_min,
            include_max=include_max,
            sort_priority=999,
            highlight_red=True,
        )
    )
    return groups


def _build_generic_transport_cost_groups(row: pd.Series) -> list[CostGroup]:
    sub_mode_raw = _normalize_text(row.get("Sub Mode")) or "UNKNOWN"
    sub_mode_normalized = _normalize_sub_mode(sub_mode_raw)
    cost_type = _normalize_text(row.get("Cost Type")) or "Transport"
    max_weight = _normalize_number(
        _value_from_aliases(
            row,
            ("Maximum weight (included)", "Maximum net weight"),
        )
    )
    min_weight = _normalize_number(
        _value_from_aliases(
            row,
            ("Minimum weight (excluded)", "Minimum weight", "Minimum weight "),
        )
    )
    min_cost = _normalize_number(row.get("Min Cost Value"))
    max_cost = _normalize_number(row.get("Max Cost Value"))

    apply_if = ""
    if max_weight is not None:
        apply_if = f"Weight/kg less than or equal to '{_format_number_text(float(max_weight))}'"
    if isinstance(min_weight, (int, float)) and float(min_weight) > 0:
        min_clause = f"Weight/kg greater than '{_format_number_text(float(min_weight))}'"
        apply_if = f"{apply_if} and {min_clause}" if apply_if else min_clause

    return [
        CostGroup(
            key=(
                f"GENERIC|{sub_mode_normalized or 'UNKNOWN'}|{cost_type.upper()}|"
                f"{max_weight if max_weight is not None else 'NA'}|{min_weight if min_weight is not None else 'NA'}|"
                f"{min_cost is not None}|{max_cost is not None}"
            ),
            cost_name=f"Transport cost ({sub_mode_raw}, {cost_type})",
            apply_if=apply_if,
            rate_by=_build_generic_rate_by(sub_mode_raw, cost_type),
            rate_types=("p/unit",),
            include_min=min_cost is not None,
            include_max=max_cost is not None,
            sort_priority=_generic_transport_sort_priority(cost_type),
        )
    ]


def _build_generic_rate_by(sub_mode: object, cost_type: str) -> str:
    base = _submode_to_rateby(sub_mode).upper()
    tokens: list[str] = []
    cost_type_upper = cost_type.upper()
    if "BARGE" in cost_type_upper:
        tokens.append("BARGE")
    if "RAIL" in cost_type_upper:
        tokens.append("RAIL")
    if "RETURN" in cost_type_upper:
        tokens.append("RETURN")
    if not tokens:
        return base
    return f"{base} {' '.join(tokens)}"


def _generic_transport_sort_priority(cost_type: str) -> int:
    text = cost_type.upper()
    has_extra_type = ("BARGE" in text) or ("RAIL" in text)
    has_return = "RETURN" in text
    if not has_extra_type and not has_return:
        return 5
    if has_return and not has_extra_type:
        return 6
    if has_extra_type and has_return:
        return 7
    return 8


def _rate_by_suffix(rate_by: str) -> str:
    text = _normalize_text(rate_by).upper()
    if text.startswith("CONTAINER/"):
        return text.split("/", 1)[1].strip()
    return text


def _base_size_from_text(value: str) -> int | None:
    match = re.search(r"(\d+)", value)
    if not match:
        return None
    size = int(match.group(1))
    return 40 if size >= 40 else 20


def _normalize_supplier_key(value: object) -> str:
    return _normalize_text(value).upper()


def _extract_first_number(value: object) -> float | None:
    number = _normalize_number(value)
    if isinstance(number, (int, float)):
        return float(number)
    text = _normalize_text(value).replace(",", ".")
    currency_match = re.search(r"(\d+(?:\.\d+)?)\s*(?:EUR|USD)", text, flags=re.IGNORECASE)
    if currency_match:
        try:
            return float(currency_match.group(1))
        except ValueError:
            pass
    match = re.search(r"(?<![A-Za-z])(\d+(?:\.\d+)?)(?![A-Za-z])", text)
    if not match:
        return None
    try:
        return float(match.group(1))
    except ValueError:
        return None


def _country_base(value: object) -> str:
    text = _normalize_text(value).upper()
    if not text:
        return ""
    return text.split("/", 1)[0].strip()


def _load_generic_services_cost_lookups(
    main_rate_selections: list[SubfolderSelection],
    *,
    imo_cost_type: str,
    include_positioning: bool,
) -> tuple[
    dict[tuple[str, int], tuple[str, float]],
    dict[int, tuple[str, float]],
    dict[tuple[str, int], tuple[str, float]],
    list[tuple[str, int, int, str | None, str, float]],
]:
    imo_lookup: dict[tuple[str, int], tuple[str, float]] = {}
    positioning_lookup: dict[int, tuple[str, float]] = {}
    t1_lookup: dict[tuple[str, int], tuple[str, float]] = {}
    waiting_entries: list[tuple[str, int, int, str | None, str, float]] = []

    for selection in main_rate_selections:
        if PREON_GENERIC_SERVICES_TAB not in selection.tabs:
            continue
        services_df = extract_sheet_to_dataframe(selection.file_path, PREON_GENERIC_SERVICES_TAB)
        services_df = services_df.rename(columns=lambda col: str(col).strip())
        required = {"Supplier", "Sub Mode", "Cost Type", "Currency", "Cost"}
        if not required.issubset(set(services_df.columns)):
            continue

        for _, row in services_df.iterrows():
            cost_type = _normalize_text(row.get("Cost Type")).upper()
            size_bucket = _base_size_from_text(_normalize_text(row.get("Sub Mode")))
            if size_bucket is None:
                continue
            currency = _normalize_text(row.get("Currency")).upper()
            cost_value = _extract_first_number(row.get("Cost"))
            if currency not in {"EUR", "USD"} or not isinstance(cost_value, (int, float)):
                continue
            numeric_cost = float(cost_value)
            supplier_key = _normalize_supplier_key(row.get("Supplier"))

            if cost_type == imo_cost_type.upper():
                if supplier_key:
                    imo_lookup[(supplier_key, size_bucket)] = (currency, numeric_cost)
            elif include_positioning and cost_type == "ALTERNATIVE CONTAINER DEPOT (<150 KM) CHARGE":
                positioning_lookup[size_bucket] = (currency, numeric_cost)
            elif cost_type == "T1 DOCUMENTATION (IF REQUIRED)":
                if supplier_key:
                    t1_lookup[(supplier_key, size_bucket)] = (currency, numeric_cost)
            elif cost_type.startswith("WAITING FEE (FREE TIME"):
                if supplier_key:
                    hour_match = re.search(r"FREE TIME\s+(\d+)\s+HRS?\)", cost_type)
                    if not hour_match:
                        continue
                    free_hours = int(hour_match.group(1))
                    country_match = re.search(r"\sIN\s([A-Z]{2})$", cost_type)
                    country_code = country_match.group(1) if country_match else None
                    waiting_entries.append(
                        (supplier_key, size_bucket, free_hours, country_code, currency, numeric_cost)
                    )

    return imo_lookup, positioning_lookup, t1_lookup, waiting_entries


def _shipment_rows_from_source_row(
    row: pd.Series,
    *,
    use_dhl_divisions_mapping: bool = False,
) -> list[dict[str, object]]:
    origin_country = _value_from_aliases(row, ("Origin Country", "Origin country"))
    origin_location_name = _value_from_aliases(
        row,
        ("Origin Location Name", "Origin location name", "Origin Name"),
    )
    origin_location = _value_from_aliases(
        row,
        ("Origin Location", "Origin location", "Origin"),
    )
    destination_country = _value_from_aliases(
        row,
        ("Destination Country", "Destination country"),
    )
    destination_location_name = _value_from_aliases(
        row,
        ("Destination Location Name", "Destination location name", "Destination Name"),
    )
    destination_location = _value_from_aliases(
        row,
        ("Destination Location", "Destination location", "Destination"),
    )
    valid_from_value = _value_from_aliases(row, ("Valid from", "Valid From"))
    valid_to_value = _value_from_aliases(row, ("Valid to", "Valid To"))

    if use_dhl_divisions_mapping:
        from_unlocode = origin_location
        to_location_or_port = destination_location
        destination_city = destination_location_name
    else:
        from_unlocode = _last_three(origin_location)
        to_location_or_port = destination_location_name
        destination_city = destination_location_name

    base = {
        "From Country": origin_country,
        "From (UN/LOCODE)": from_unlocode,
        "From(Location/Port)": origin_location_name,
        "Origin city": origin_location_name,
        "To Country": destination_country,
        "To (UN/LOCODE)": _last_three(destination_location),
        "to (Location/Port)": to_location_or_port,
        "Destination city": destination_city,
        "Valid From": _format_date(valid_from_value),
        "Valid To": _format_date(valid_to_value),
    }
    pre = dict(base)
    pre["SERVICE"] = "PRE-CARRIAGE"
    pre["From (UN/LOCODE)"] = None
    pre["Destination city"] = None

    on = dict(base)
    on["SERVICE"] = "ONCARRIAGE"
    on["Origin city"] = None
    on["To (UN/LOCODE)"] = None
    return [pre, on]


def _generic_shipment_columns_from_df(df: pd.DataFrame) -> list[str]:
    columns = [str(col).strip() for col in df.columns]
    lower = {name.lower(): idx for idx, name in enumerate(columns)}
    cost_type_idx = lower.get("cost type")
    if cost_type_idx is None:
        cost_type_idx = lower.get("cost type")
    if cost_type_idx is None:
        return []
    shipment_columns = columns[:cost_type_idx]
    excluded = {
        "Type Mode",
        "Sub Mode",
        "Origin Location Type",
        "Destination Location Name",
        "Destination Location Type",
    }
    filtered_columns = [col for col in shipment_columns if col not in excluded]
    return ["Forwarded" if col == "Supplier" else col for col in filtered_columns]


def _shipment_row_from_source_row_generic(
    row: pd.Series,
    shipment_columns: list[str],
) -> dict[str, object]:
    values: dict[str, object] = {}
    for col in shipment_columns:
        source_col = "Supplier" if col == "Forwarded" else col
        value = row.get(source_col)
        values[col] = value
    return values


def _generic_transport_cost_from_row(row: pd.Series) -> object:
    # Haulage sources can use "Rate"/"Rates" while other generic sources use "Cost".
    if "Rate" in row.index:
        value = _normalize_number(row.get("Rate"))
        if value is not None:
            return value
    if "Rates" in row.index:
        value = _normalize_number(row.get("Rates"))
        if value is not None:
            return value
    return _normalize_number(row.get("Cost"))


def _use_dhl_divisions_mapping(shipper: str, selection: SubfolderSelection) -> bool:
    if shipper != "Siemens Divisions":
        return False
    carrier_code = carrier_code_from_filename(
        selection.file_path.name,
        shipper=shipper,
        flow="FCL",
    )
    return carrier_code == "DHLGLOB-DE-FFM-00"


def resolve_preon_carrier_slug(
    shipper: str,
    individual_selections: list[SubfolderSelection],
) -> str:
    carrier_codes: list[str] = []
    carrier_keys: list[str] = []
    for selection in individual_selections:
        carrier_code = carrier_code_from_filename(
            selection.file_path.name,
            shipper=shipper,
            flow="FCL",
        )
        if carrier_code:
            carrier_codes.append(carrier_code)
            continue
        carrier_key = detect_carrier_key(
            selection.file_path.name,
            shipper=shipper,
            flow="FCL",
        )
        if carrier_key:
            carrier_keys.append(carrier_key)

    if carrier_codes:
        return slugify("_".join(sorted(set(carrier_codes))))
    if carrier_keys:
        return slugify("_".join(sorted(set(carrier_keys))))
    return "unknown"


def build_output_rate_card_path(flow: str, shipper: str, carrier_slug: str | None = None) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    shipper_slug = slugify(shipper)
    if carrier_slug:
        return OUTPUT_DIR / f"{flow}_{shipper_slug}_{carrier_slug}_rate_card.xlsx"
    return OUTPUT_DIR / f"{flow}_{shipper_slug}_rate_card.xlsx"


def save_preon_per_carrier_rate_card(
    shipper: str,
    individual_selections: list[SubfolderSelection],
    output_path: Path | None = None,
) -> tuple[Path, pd.DataFrame]:
    rows_by_key: dict[tuple, dict[str, object]] = {}
    cost_defs: dict[str, CostGroup] = {}
    costs_by_row: dict[tuple, dict[str, tuple[object, object, object, object]]] = {}

    for selection in individual_selections:
        if PREON_TAB not in selection.tabs:
            continue
        df = pd.read_excel(selection.file_path, sheet_name=PREON_TAB, header=3)
        df = df.rename(columns=lambda col: str(col).strip())
        for _, src_row in df.iterrows():
            groups = _build_cost_groups(src_row)
            if not groups:
                continue
            currency = src_row.get("Currency")
            cost = _generic_transport_cost_from_row(src_row)
            min_cost = _normalize_number(src_row.get("Min Cost Value"))
            max_cost = _normalize_number(src_row.get("Max Cost Value"))

            for group in groups:
                cost_defs[group.key] = group
                for shipment in _shipment_rows_from_source_row(
                    src_row,
                    use_dhl_divisions_mapping=_use_dhl_divisions_mapping(shipper, selection),
                ):
                    key = tuple(shipment[col] for col in SHIPMENT_COLUMNS)
                    rows_by_key.setdefault(key, shipment)
                    costs_by_row.setdefault(key, {})
                    costs_by_row[key][group.key] = (
                        None if pd.isna(currency) else str(currency).strip().upper(),
                        min_cost,
                        cost,
                        max_cost,
                    )

    if not rows_by_key:
        raise ValueError("No Pre-On-Carriage_RoW rows found in selected individual rate files.")

    sorted_row_keys = sorted(
        rows_by_key.keys(),
        key=_row_sort_key,
    )
    shipment_rows = [rows_by_key[key] for key in sorted_row_keys]
    shipment_df = pd.DataFrame(shipment_rows, columns=SHIPMENT_COLUMNS)

    ordered_groups = sorted(
        cost_defs.values(),
        key=_cost_group_sort_key,
    )
    cost_blocks: list[pd.DataFrame] = []
    rendered_groups: list[CostGroup] = []
    for group in ordered_groups:
        currencies: list[object] = []
        mins: list[object] = []
        rates: list[object] = []
        maxes: list[object] = []
        for key in sorted_row_keys:
            ccy, min_val, rate_val, max_val = costs_by_row.get(key, {}).get(group.key, (None, None, None, None))
            currencies.append(ccy)
            if group.include_min:
                mins.append(min_val)
            rates.append(rate_val)
            if group.include_max:
                maxes.append(max_val)

        has_any_value = any(value is not None for value in currencies) or any(
            value is not None for value in rates
        )
        if group.include_min:
            has_any_value = has_any_value or any(value is not None for value in mins)
        if group.include_max:
            has_any_value = has_any_value or any(value is not None for value in maxes)
        if not has_any_value:
            continue

        block = {f"{group.key}__currency": currencies}
        if group.include_min:
            block[f"{group.key}__min"] = mins
        block[f"{group.key}__rate"] = rates
        if group.include_max:
            block[f"{group.key}__max"] = maxes
        cost_blocks.append(pd.DataFrame(block))
        rendered_groups.append(group)

    rate_card_df = pd.concat([shipment_df, *cost_blocks], axis=1)

    carrier_slug = resolve_preon_carrier_slug(shipper, individual_selections)
    if output_path is None:
        output_path = build_output_rate_card_path("Pre_on_carriage", shipper, carrier_slug)
    workbook = Workbook()
    _write_sheet(workbook, rate_card_df, rendered_groups)
    transport_cost_names = list(
        dict.fromkeys(
            group.cost_name for group in rendered_groups if group.cost_name.startswith("Transport cost")
        )
    )
    accessorial_df = build_preon_accessorial_costs_dataframe(
        shipper,
        individual_selections,
        transport_cost_names,
    )
    _write_accessorial_sheet(workbook, accessorial_df)
    workbook.save(output_path)
    return output_path, rate_card_df


def _normalized_thc_container_label(sub_mode: object) -> str:
    label = _normalize_sub_mode(sub_mode)
    if label:
        return label
    return "20FT"


def _thc_lookup_code(sub_mode: object) -> str:
    text = _normalize_text(sub_mode).upper()
    match = re.search(r"(\d+)", text)
    if not match:
        return "20_STD"
    size = int(match.group(1))
    normalized_size = 40 if size >= 40 else 20
    return f"{normalized_size}_STD"


def _load_generic_thc_from_digi_fcl_rates(
    main_rate_selections: list[SubfolderSelection],
) -> dict[int, tuple[str, float]]:
    for selection in main_rate_selections:
        if "DIGI_FCL_Rates" not in selection.tabs:
            continue
        df = extract_sheet_to_dataframe(selection.file_path, "DIGI_FCL_Rates")
        df = df.rename(columns=lambda col: str(col).strip())
        required_columns = {
            "Origin Country",
            "Container Size",
            "THC indication Origin Currency",
            "THC indication Origin lump sum",
        }
        if not required_columns.issubset(set(df.columns)):
            continue

        working = df.copy()
        working["Origin Country"] = working["Origin Country"].astype(str).str.strip().str.upper()
        working = working[working["Origin Country"] == "DE"]
        if working.empty:
            continue

        values: dict[int, tuple[str, float]] = {}
        for _, row in working.iterrows():
            size_num = _normalize_number(row.get("Container Size"))
            if not isinstance(size_num, (int, float)):
                continue
            size_bucket = 40 if int(size_num) >= 40 else 20
            if size_bucket not in {20, 40}:
                continue

            lump_sum = _normalize_number(row.get("THC indication Origin lump sum"))
            if not isinstance(lump_sum, (int, float)):
                continue
            currency = _normalize_text(row.get("THC indication Origin Currency")).upper()
            if currency not in {"EUR", "USD"}:
                continue
            if size_bucket not in values:
                values[size_bucket] = (currency, float(lump_sum))

        if values:
            return values
    return {}


def save_preon_generic_rate_card(
    shipper: str,
    main_rate_selections: list[SubfolderSelection],
    output_path: Path | None = None,
    thc_lookup: FclThcLookup | None = None,
    *,
    source_tab: str = PREON_GENERIC_TAB,
    include_thc_origin: bool = True,
    include_positioning: bool = True,
    include_terms_accessorial: bool = True,
    include_add_services_accessorial: bool = True,
    services_imo_cost_type: str = "IMO charge",
) -> tuple[Path, pd.DataFrame]:
    shipment_columns: list[str] | None = None
    rows_by_key: dict[tuple, dict[str, object]] = {}
    cost_defs: dict[str, CostGroup] = {}
    costs_by_row: dict[tuple, dict[str, tuple[object, object, object, object]]] = {}
    thc_source_by_row: dict[tuple, tuple[object, object, object, object]] = {}
    row_rate_by_suffixes: dict[tuple, set[str]] = {}

    for selection in main_rate_selections:
        if source_tab not in selection.tabs:
            continue
        df = extract_sheet_to_dataframe(selection.file_path, source_tab)
        df = df.rename(columns=lambda col: str(col).strip())
        if shipment_columns is None:
            shipment_columns = _generic_shipment_columns_from_df(df)
            if not shipment_columns:
                raise ValueError(
                    f"{source_tab} is missing 'Cost Type' column needed for generic shipment columns."
                )
        for _, src_row in df.iterrows():
            groups = _build_generic_transport_cost_groups(src_row)
            if not groups:
                continue
            currency = src_row.get("Currency")
            cost = _generic_transport_cost_from_row(src_row)
            min_cost = _normalize_number(src_row.get("Min Cost Value"))
            max_cost = _normalize_number(src_row.get("Max Cost Value"))

            for group in groups:
                cost_defs[group.key] = group
                shipment = _shipment_row_from_source_row_generic(src_row, shipment_columns)
                key = tuple(shipment[col] for col in shipment_columns)
                rows_by_key.setdefault(key, shipment)
                costs_by_row.setdefault(key, {})
                costs_by_row[key][group.key] = (
                    None if pd.isna(currency) else str(currency).strip().upper(),
                    min_cost,
                    cost,
                    max_cost,
                )
                row_rate_by_suffixes.setdefault(key, set()).add(_rate_by_suffix(group.rate_by))
                thc_source_by_row[key] = (
                    src_row.get("Origin Country"),
                    src_row.get("Destination Country"),
                    src_row.get("Sub Mode"),
                    currency,
                )

    if not rows_by_key:
        raise ValueError(f"No {source_tab} rows found in selected main rates files.")

    thc_origin_values: dict[int, tuple[str, float]] = {}
    if include_thc_origin:
        thc_origin_values = _load_generic_thc_from_digi_fcl_rates(main_rate_selections)
        if not thc_origin_values and thc_lookup is None:
            try:
                thc_lookup = load_fcl_thc_lookup()
            except (FileNotFoundError, KeyError):
                thc_lookup = None

    services_imo_lookup: dict[tuple[str, int], tuple[str, float]] = {}
    services_positioning_lookup: dict[int, tuple[str, float]] = {}
    services_t1_lookup: dict[tuple[str, int], tuple[str, float]] = {}
    services_waiting_entries: list[tuple[str, int, int, str | None, str, float]] = []
    if shipper == "Siemens Healthineers":
        (
            services_imo_lookup,
            services_positioning_lookup,
            services_t1_lookup,
            services_waiting_entries,
        ) = _load_generic_services_cost_lookups(
            main_rate_selections,
            imo_cost_type=services_imo_cost_type,
            include_positioning=include_positioning,
        )

    if (
        (include_thc_origin and (thc_origin_values or thc_lookup is not None))
        or services_imo_lookup
        or services_positioning_lookup
        or services_t1_lookup
        or services_waiting_entries
    ):
        thc_defs: dict[str, CostGroup] = {}
        unique_rate_by_suffixes = sorted(
            {
                suffix
                for suffixes in row_rate_by_suffixes.values()
                for suffix in suffixes
                if suffix
            }
        )
        for suffix in unique_rate_by_suffixes:
            thc_key = f"THCORIGIN|{suffix}"
            if include_thc_origin:
                thc_defs[thc_key] = CostGroup(
                    key=thc_key,
                    cost_name=f"THC Origin ({suffix})",
                    apply_if="",
                    rate_by=f"Container/{suffix}",
                    rate_types=("p/unit",),
                    include_min=False,
                    include_max=False,
                    sort_priority=60,
                )
            if services_imo_lookup:
                imo_key = f"IMO|{suffix}"
                thc_defs[imo_key] = CostGroup(
                    key=imo_key,
                    cost_name=f"IMO Charges ({suffix})",
                    apply_if="",
                    rate_by=f"Container/{suffix}",
                    rate_types=("p/unit",),
                    include_min=False,
                    include_max=False,
                    sort_priority=70,
                )
            if services_positioning_lookup:
                positioning_key = f"CONTPOS|{suffix}"
                thc_defs[positioning_key] = CostGroup(
                    key=positioning_key,
                    cost_name=f"Container positioning ({suffix})",
                    apply_if="",
                    rate_by=f"Container/{suffix}",
                    rate_types=("p/unit",),
                    include_min=False,
                    include_max=False,
                    sort_priority=71,
                )
            if services_waiting_entries:
                continue

        if services_t1_lookup:
            thc_defs["T1DOC|CEVA"] = CostGroup(
                key="T1DOC|CEVA",
                cost_name="T1 Document Fee (CEVA)",
                apply_if="",
                rate_by="per shipment",
                rate_types=("p/unit",),
                include_min=False,
                include_max=False,
                sort_priority=72,
            )
        if services_waiting_entries:
            waiting_hours = sorted({hours for _, _, hours, _, _, _ in services_waiting_entries})
            for hours in waiting_hours:
                wait_key = f"WAIT{hours}H"
                thc_defs[wait_key] = CostGroup(
                    key=wait_key,
                    cost_name=f"Waiting fee (free time {hours} hrs)",
                    apply_if="",
                    rate_by="per shipment",
                    rate_types=("p/unit",),
                    include_min=False,
                    include_max=False,
                    sort_priority=73,
                )
            thc_defs["T1DOC|DHL"] = CostGroup(
                key="T1DOC|DHL",
                cost_name="T1 Document Fee (DHL)",
                apply_if="",
                rate_by="per shipment",
                rate_types=("p/unit",),
                include_min=False,
                include_max=False,
                sort_priority=72,
            )

        for key in rows_by_key:
            origin_country, destination_country, sub_mode, currency = thc_source_by_row.get(
                key,
                (None, None, None, None),
            )
            currency_text = None if pd.isna(currency) else str(currency).strip().upper()
            if currency_text not in {"EUR", "USD"}:
                continue
            lookup_code = _thc_lookup_code(sub_mode)
            inbound: float | None = None
            inbound_currency: str | None = currency_text
            size_bucket = 40 if lookup_code.startswith("40_") else 20
            if include_thc_origin:
                origin_entry = thc_origin_values.get(size_bucket)
                if origin_entry is not None:
                    origin_currency, origin_value = origin_entry
                    inbound = origin_value
                    inbound_currency = origin_currency
                elif thc_lookup is not None:
                    inbound = thc_lookup.lookup(origin_country, lookup_code, currency_text)
            row_costs = costs_by_row.setdefault(key, {})
            if include_thc_origin and inbound is not None:
                for suffix in row_rate_by_suffixes.get(key, set()):
                    suffix_base_size = _base_size_from_text(suffix)
                    if suffix_base_size is None or suffix_base_size != size_bucket:
                        continue
                    thc_key = f"THCORIGIN|{suffix}"
                    row_costs[thc_key] = (inbound_currency, None, inbound, None)

            shipment = rows_by_key.get(key, {})
            forwarder_key = _normalize_supplier_key(shipment.get("Forwarded"))
            lane_origin = _country_base(shipment.get("Origin Country"))
            lane_destination = _country_base(shipment.get("Destination Country"))
            for suffix in row_rate_by_suffixes.get(key, set()):
                suffix_base_size = _base_size_from_text(suffix)
                if suffix_base_size is None:
                    continue
                imo_entry = services_imo_lookup.get((forwarder_key, suffix_base_size))
                if imo_entry is not None:
                    imo_currency, imo_cost = imo_entry
                    row_costs[f"IMO|{suffix}"] = (imo_currency, None, imo_cost, None)
                positioning_entry = services_positioning_lookup.get(suffix_base_size)
                if positioning_entry is not None:
                    pos_currency, pos_cost = positioning_entry
                    row_costs[f"CONTPOS|{suffix}"] = (pos_currency, None, pos_cost, None)
                t1_entry = services_t1_lookup.get((forwarder_key, suffix_base_size))
                if t1_entry is not None:
                    t1_currency, t1_cost = t1_entry
                    row_costs["T1DOC|CEVA"] = (t1_currency, None, t1_cost, None)
                    row_costs["T1DOC|DHL"] = (t1_currency, None, t1_cost, None)
                waiting_match: tuple[str, float] | None = None
                waiting_by_hours: dict[int, tuple[str, float]] = {}
                for supplier, size, hours, country_code, wait_currency, wait_cost in services_waiting_entries:
                    if supplier != forwarder_key or size != suffix_base_size:
                        continue
                    if country_code is None:
                        waiting_by_hours.setdefault(hours, (wait_currency, wait_cost))
                        continue
                    if lane_origin == country_code or lane_destination == country_code:
                        waiting_by_hours[hours] = (wait_currency, wait_cost)
                for hours, waiting_match in waiting_by_hours.items():
                    wait_currency, wait_cost = waiting_match
                    row_costs[f"WAIT{hours}H"] = (wait_currency, None, wait_cost, None)
        cost_defs.update(thc_defs)

    assert shipment_columns is not None
    sorted_row_keys = sorted(
        rows_by_key.keys(),
        key=lambda key: tuple("" if value is None else str(value) for value in key),
    )
    shipment_rows = [rows_by_key[key] for key in sorted_row_keys]
    shipment_df = pd.DataFrame(shipment_rows, columns=shipment_columns)

    ordered_groups = sorted(cost_defs.values(), key=_cost_group_sort_key)
    cost_blocks: list[pd.DataFrame] = []
    rendered_groups: list[CostGroup] = []
    for group in ordered_groups:
        currencies: list[object] = []
        mins: list[object] = []
        rates: list[object] = []
        maxes: list[object] = []
        for key in sorted_row_keys:
            ccy, min_val, rate_val, max_val = costs_by_row.get(key, {}).get(group.key, (None, None, None, None))
            currencies.append(ccy)
            if group.include_min:
                mins.append(min_val)
            rates.append(rate_val)
            if group.include_max:
                maxes.append(max_val)

        has_any_value = any(value is not None for value in currencies) or any(
            value is not None for value in rates
        )
        if group.include_min:
            has_any_value = has_any_value or any(value is not None for value in mins)
        if group.include_max:
            has_any_value = has_any_value or any(value is not None for value in maxes)
        if not has_any_value:
            continue

        block = {f"{group.key}__currency": currencies}
        if group.include_min:
            block[f"{group.key}__min"] = mins
        block[f"{group.key}__rate"] = rates
        if group.include_max:
            block[f"{group.key}__max"] = maxes
        cost_blocks.append(pd.DataFrame(block))
        rendered_groups.append(group)

    rate_card_df = pd.concat([shipment_df, *cost_blocks], axis=1)

    if output_path is None:
        output_path = build_output_rate_card_path("Pre_on_carriage_generic", shipper)
    workbook = Workbook()
    _write_sheet(
        workbook,
        rate_card_df,
        rendered_groups,
        shipment_columns=shipment_columns,
        bold_headers=GENERIC_BOLD_SHIPMENT_HEADERS,
    )
    unique_rate_by_suffixes = sorted(
        {
            suffix
            for suffixes in row_rate_by_suffixes.values()
            for suffix in suffixes
            if suffix
        }
    )
    generic_accessorial_df = build_preon_generic_accessorial_costs_dataframe(
        main_rate_selections,
        shipper=shipper,
        rate_by_suffixes=unique_rate_by_suffixes,
        include_terms_costs=include_terms_accessorial,
        include_add_services_costs=include_add_services_accessorial,
    )
    _write_accessorial_sheet(workbook, generic_accessorial_df)
    workbook.save(output_path)
    return output_path, rate_card_df


def _row_sort_key(key: tuple) -> tuple:
    service = key[SHIPMENT_COLUMNS.index("SERVICE")]
    normalized = tuple("" if value is None else str(value) for value in key)
    return (SERVICE_ORDER.get(str(service), 99),) + normalized


def _row_sort_key_by_columns(key: tuple, shipment_columns: list[str]) -> tuple:
    service = key[shipment_columns.index("SERVICE")]
    normalized = tuple("" if value is None else str(value) for value in key)
    return (SERVICE_ORDER.get(str(service), 99),) + normalized


def _cost_group_sort_key(group: CostGroup) -> tuple:
    type_mode = group.key.split("|", 1)[0].upper()
    return (group.sort_priority, TYPE_MODE_ORDER.get(type_mode, 99), group.cost_name, group.key)


def _dieselfloater_cost_text(raw_value: object) -> tuple[str, bool]:
    if raw_value is None or (isinstance(raw_value, float) and pd.isna(raw_value)):
        return "", False

    if isinstance(raw_value, str):
        value = raw_value.strip()
        if not value:
            return "", False
        if "%" in value:
            return value, True
        numeric = _normalize_number(value)
        if isinstance(numeric, (int, float)):
            if 0 <= float(numeric) <= 1:
                return f"{_format_number_text(float(numeric) * 100)}%", True
            return _format_number_text(float(numeric)), False
        return value, False

    numeric = _normalize_number(raw_value)
    if isinstance(numeric, (int, float)):
        if 0 <= float(numeric) <= 1:
            return f"{_format_number_text(float(numeric) * 100)}%", True
        return _format_number_text(float(numeric)), False
    return str(raw_value), False


def _parse_amount_text(value: str) -> float | None:
    cleaned = value.strip().replace(",", ".")
    cleaned = re.sub(r"[^0-9.]", "", cleaned)
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _format_tmp_price(unit_rate: float | None, max_cap: float | None = None) -> str:
    if unit_rate is None:
        return ""
    if max_cap is None:
        return f"EUR {_format_number_text(unit_rate)}"
    return f"EUR {_format_number_text(unit_rate)} (MAX {_format_number_text(max_cap)})"


def _extract_dhl_tmp_rows_from_glossary(source_file: Path) -> list[dict[str, str]]:
    glossary_df = extract_sheet_to_dataframe(source_file, GLOSSARY_TAB)
    tmp_text = _glossary_fee_text(glossary_df, "TMP-Fee-ALL_IN")
    if not tmp_text:
        return []

    dhl_section = tmp_text
    marker = re.search(r"Siemens Healthineers", tmp_text, flags=re.IGNORECASE)
    if marker:
        dhl_section = tmp_text[: marker.start()]

    rows: list[dict[str, str]] = []

    fcl_match = re.search(
        r"TMP Fee worldwide[\s\S]*?EUR\s*([\d,\.]+)[\s\S]*?CAP\)?\s*EUR\s*([\d,\.]+)",
        dhl_section,
        flags=re.IGNORECASE,
    )
    if fcl_match:
        unit = _parse_amount_text(fcl_match.group(1))
        cap = _parse_amount_text(fcl_match.group(2))
        rows.append(
            {
                "Cost Name": "TMP Fee(TMP-Fee-ALL_IN, FCL)",
                "Rate by": "Quantity/Container",
                "Apply over": "",
                "Cost": _format_tmp_price(unit, cap),
                "Valid From": "",
                "Valid To": "",
            }
        )

    addon_match = re.search(
        r"Per delivery note:\s*EUR\s*([\d,\.]+)[\s\S]*?CAP.*?EUR\s*([\d,\.]+)",
        dhl_section,
        flags=re.IGNORECASE,
    )
    if addon_match:
        unit = _parse_amount_text(addon_match.group(1))
        cap = _parse_amount_text(addon_match.group(2))
        rows.append(
            {
                "Cost Name": "TMP Fee(TMP-Fee-ALL_IN, Consol)",
                "Rate by": "per delivery note",
                "Apply over": "",
                "Cost": _format_tmp_price(unit, cap),
                "Valid From": "",
                "Valid To": "",
            }
        )

    lcl_match = re.search(r"LCL Fee:\s*EUR\s*([\d,\.]+)", dhl_section, flags=re.IGNORECASE)
    if lcl_match:
        unit = _parse_amount_text(lcl_match.group(1))
        rows.append(
            {
                "Cost Name": "TMP Fee(TMP-Fee-ALL_IN, LCL)",
                "Rate by": "per House B/L",
                "Apply over": "",
                "Cost": _format_tmp_price(unit),
                "Valid From": "",
                "Valid To": "",
            }
        )

    return rows


def build_preon_generic_accessorial_costs_dataframe(
    main_rate_selections: list[SubfolderSelection],
    *,
    shipper: str,
    rate_by_suffixes: list[str],
    include_terms_costs: bool = True,
    include_add_services_costs: bool = True,
) -> pd.DataFrame:
    rows: list[dict[str, str]] = []
    if include_add_services_costs:
        for selection in main_rate_selections:
            tab_name = next((name for name in PREON_GENERIC_ADD_SERVICES_TABS if name in selection.tabs), None)
            if tab_name is None:
                continue
            try:
                add_services_df = pd.read_excel(selection.file_path, sheet_name=tab_name, header=1)
            except ValueError:
                continue
            add_services_df = add_services_df.rename(columns=lambda col: str(col).strip())
            required = {"Scope", "Currency", "Rate", "Unit", "Valid from", "Valid until"}
            if not required.issubset(set(add_services_df.columns)):
                continue

            for _, row in add_services_df.iterrows():
                scope = _normalize_text(row.get("Scope"))
                if not scope:
                    continue
                currency = _normalize_text(row.get("Currency")).upper()
                rate_value = _normalize_number(row.get("Rate"))
                unit = _normalize_text(row.get("Unit"))
                if rate_value is None and not currency:
                    continue
                cost_text = ""
                if isinstance(rate_value, (int, float)):
                    if currency:
                        cost_text = f"{currency} {_format_number_text(float(rate_value))}"
                    else:
                        cost_text = _format_number_text(float(rate_value))
                else:
                    cost_text = _normalize_text(row.get("Rate"))
                    if currency and cost_text:
                        cost_text = f"{currency} {cost_text}"
                rows.append(
                    {
                        "Cost Name": scope,
                        "Rate by": unit,
                        "Apply over": "",
                        "Cost": cost_text,
                        "Valid From": _format_date(row.get("Valid from")) or "",
                        "Valid To": _format_date(row.get("Valid until")) or "",
                    }
                )

    if not rows:
        base_df = pd.DataFrame(columns=PREON_ACCESSORIAL_COLUMNS)
    else:
        base_df = pd.DataFrame(rows, columns=PREON_ACCESSORIAL_COLUMNS)

    if not include_terms_costs or shipper != "Siemens Healthineers" or not rate_by_suffixes:
        return base_df

    terms_rows = _build_healthineers_terms_accessorial_rows(main_rate_selections, rate_by_suffixes)
    if not terms_rows:
        return base_df
    terms_df = pd.DataFrame(terms_rows, columns=PREON_ACCESSORIAL_COLUMNS)
    if base_df.empty:
        return terms_df
    return pd.concat([base_df, terms_df], ignore_index=True)


def build_preon_accessorial_costs_dataframe(
    shipper: str,
    individual_selections: list[SubfolderSelection],
    transport_cost_names: list[str],
) -> pd.DataFrame:
    rows: list[dict[str, str]] = []
    apply_over_text = "; ".join(transport_cost_names)

    for selection in individual_selections:
        if (
            shipper == "Siemens Divisions"
            and GLOSSARY_TAB in selection.tabs
            and carrier_code_from_filename(selection.file_path.name, shipper=shipper, flow="FCL")
            == "DHLGLOB-DE-FFM-00"
        ):
            rows.extend(_extract_dhl_tmp_rows_from_glossary(selection.file_path))

        if PREON_DIESELFLOATER_TAB not in selection.tabs:
            continue

        dieselfloater_df = extract_sheet_to_dataframe(selection.file_path, PREON_DIESELFLOATER_TAB)
        for _, row in dieselfloater_df.iterrows():
            cost_text, is_percent = _dieselfloater_cost_text(row.get("Dieselfloater Add-on"))
            if not cost_text:
                continue

            rows.append(
                {
                    "Cost Name": "Emergency Fuel Surcharge",
                    "Rate by": "per shipment",
                    "Apply over": apply_over_text if is_percent else "",
                    "Cost": cost_text,
                    "Valid From": _format_date(row.get("Valid from")) or "",
                    "Valid To": _format_date(row.get("Valid until")) or "",
                }
            )

    if not rows:
        return pd.DataFrame(columns=PREON_ACCESSORIAL_COLUMNS)
    return pd.DataFrame(rows, columns=PREON_ACCESSORIAL_COLUMNS)


def _terms_table_rows(file_path: Path, tab_name: str, short_name: str) -> list[dict[str, object]]:
    raw = pd.read_excel(file_path, sheet_name=tab_name, header=None)
    marker = f"{short_name} - Agreed Pick-up and Drop-off fees Germany"
    marker_idx: int | None = None
    for idx in range(len(raw)):
        row_text = " ".join(_normalize_text(value) for value in raw.iloc[idx].tolist())
        if marker in row_text:
            marker_idx = idx
            break
    if marker_idx is None:
        return []

    depot_header_idx = marker_idx + 3
    if depot_header_idx >= len(raw):
        return []
    header_row = raw.iloc[depot_header_idx]
    depot_col = None
    for col_idx, value in enumerate(header_row.tolist()):
        if _normalize_text(value).lower() == "depot":
            depot_col = col_idx
            break
    if depot_col is None:
        return []

    rows: list[dict[str, object]] = []
    for row_idx in range(depot_header_idx + 1, len(raw)):
        depot = _normalize_text(raw.iat[row_idx, depot_col])
        if not depot:
            break
        rows.append(
            {
                "Depot": depot,
                "PICK20": _normalize_number(raw.iat[row_idx, depot_col + 1]),
                "PICK40": _normalize_number(raw.iat[row_idx, depot_col + 2]),
                "PICK40HC": _normalize_number(raw.iat[row_idx, depot_col + 3]),
                "DROP20": _normalize_number(raw.iat[row_idx, depot_col + 4]),
                "DROP40": _normalize_number(raw.iat[row_idx, depot_col + 5]),
                "DROP40HC": _normalize_number(raw.iat[row_idx, depot_col + 6]),
            }
        )
    return rows


def _carrier_condition_for_cost(cost_kind: str) -> str:
    if cost_kind == "empty":
        return "CARRIER_NAME equals 'DHL GLO-DE-FRAN-512'"
    if cost_kind == "pre":
        return (
            "CARRIER_NAME equals 'KUEHNE-DE-DUES-462' or "
            "CARRIER_NAME equals 'KUEHNE-DE-FREN-464' or "
            "CARRIER_NAME equals 'KUEHNE-DE-NUER-103' or "
            "CARRIER_NAME equals 'KUEHNE-DE-BREM-694'"
        )
    return "CARRIER_NAME equals 'CEVA AI-DE-FRAN-281'"


def _pick_drop_value(table_row: dict[str, object], suffix: str, *, pick: bool) -> float | None:
    is_40hc = "40HC" in suffix
    base_size = _base_size_from_text(suffix)
    if base_size == 20:
        key = "PICK20" if pick else "DROP20"
    elif is_40hc:
        key = "PICK40HC" if pick else "DROP40HC"
    else:
        key = "PICK40" if pick else "DROP40"
    value = table_row.get(key)
    if isinstance(value, (int, float)):
        return float(value)
    return None


def _build_healthineers_terms_accessorial_rows(
    main_rate_selections: list[SubfolderSelection],
    rate_by_suffixes: list[str],
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []

    for selection in main_rate_selections:
        for tab_name, short_name in PREON_GENERIC_TERMS_TAB_SHORT.items():
            if tab_name not in selection.tabs:
                continue
            table_rows = _terms_table_rows(selection.file_path, tab_name, short_name)
            if not table_rows:
                continue

            for suffix in rate_by_suffixes:
                suffix_upper = suffix.upper()
                is_return = "RETURN" in suffix_upper
                suffix_label = suffix_upper.replace(" RETURN", "")
                return_tail = " - RETURN" if is_return else ""
                for table_row in table_rows:
                    depot = _normalize_text(table_row.get("Depot"))
                    if not depot:
                        continue
                    pick_value = _pick_drop_value(table_row, suffix_upper, pick=True)
                    drop_value = _pick_drop_value(table_row, suffix_upper, pick=False)

                    def build_apply_if(cost_kind: str) -> str:
                        return (
                            f"SCAS equals '{short_name}' and "
                            f"CONTAINER_DEPOT equals '{depot}' and "
                            f"{_carrier_condition_for_cost(cost_kind)}"
                        )

                    if pick_value is not None:
                        rows.append(
                            {
                                "Cost Name": f"Empty Container Fee ({suffix_label} {short_name}{return_tail})",
                                "Rate by": f"Container/{suffix_upper}",
                                "Apply over": build_apply_if("empty"),
                                "Cost": f"EUR {_format_number_text(pick_value)}",
                                "Valid From": "",
                                "Valid To": "",
                            }
                        )
                        rows.append(
                            {
                                "Cost Name": (
                                    f"Pre-Carriage ({short_name} - Agreed pick-up fees "
                                    f"{suffix_label}{return_tail})"
                                ),
                                "Rate by": f"Container/{suffix_upper}",
                                "Apply over": build_apply_if("pre"),
                                "Cost": f"EUR {_format_number_text(pick_value)}",
                                "Valid From": "",
                                "Valid To": "",
                            }
                        )

                    if short_name == "ONEY":
                        if pick_value is not None:
                            rows.append(
                                {
                                    "Cost Name": (
                                        f"Pickup Fee ({short_name} - Agreed pick-up fees "
                                        f"{suffix_label}{return_tail})"
                                    ),
                                    "Rate by": f"Container/{suffix_upper}",
                                    "Apply over": build_apply_if("oney"),
                                    "Cost": f"EUR {_format_number_text(pick_value)}",
                                    "Valid From": "",
                                    "Valid To": "",
                                }
                            )
                        if drop_value is not None:
                            rows.append(
                                {
                                    "Cost Name": (
                                        f"Drop Off Fee ({short_name} - Agreed drop-off fees "
                                        f"{suffix_label}{return_tail})"
                                    ),
                                    "Rate by": f"Container/{suffix_upper}",
                                    "Apply over": build_apply_if("oney"),
                                    "Cost": f"EUR {_format_number_text(drop_value)}",
                                    "Valid From": "",
                                    "Valid To": "",
                                }
                            )
    return rows


def _write_sheet(
    workbook: Workbook,
    df: pd.DataFrame,
    groups: list[CostGroup],
    *,
    shipment_columns: list[str] | None = None,
    bold_headers: set[str] | None = None,
) -> None:
    ws = workbook.active
    ws.title = RATE_CARD_SHEET_NAME
    if shipment_columns is None:
        shipment_columns = SHIPMENT_COLUMNS
    if bold_headers is None:
        bold_headers = BOLD_SHIPMENT_HEADERS
    shipment_count = len(shipment_columns)
    header_rows = 4
    next_col = shipment_count + 1

    group_ranges: list[tuple[int, int]] = []
    for g in groups:
        rate_label = g.rate_types[0]
        count = 2 + int(g.include_min) + int(g.include_max)
        start = next_col
        end = start + count - 1
        group_ranges.append((start, end))

        for row in (1, 2, 3):
            ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
        ws.cell(1, start, g.cost_name)
        ws.cell(2, start, g.apply_if)
        ws.cell(3, start, g.rate_by)
        ws.cell(4, start, "Currency")
        col = start + 1
        if g.include_min:
            ws.cell(4, col, "MIN Flat")
            col += 1
        ws.cell(4, col, rate_label)
        col += 1
        if g.include_max:
            ws.cell(4, col, "MAX Flat")

        if g.highlight_red:
            red_font = Font(color="FF0000", bold=True, size=10)
            for row in (1, 2, 3, 4):
                for col_idx in range(start, end + 1):
                    ws.cell(row, col_idx).font = red_font
        next_col = end + 1

    for idx, name in enumerate(shipment_columns, start=1):
        ws.cell(4, idx, name)

    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=5):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row_idx, col_idx, value)
            if isinstance(value, (int, float)):
                cell.number_format = _number_format(value)

    _format_sheet(
        ws,
        shipment_count,
        next_col - 1,
        5,
        4 + len(df),
        group_ranges,
        shipment_columns=shipment_columns,
        bold_headers=bold_headers,
    )


def _write_accessorial_sheet(workbook: Workbook, accessorial_df: pd.DataFrame) -> None:
    ws = workbook.create_sheet(title=ACCESSORIAL_COSTS_SHEET_NAME)
    if accessorial_df.empty:
        for idx, col in enumerate(PREON_ACCESSORIAL_COLUMNS, start=1):
            ws.cell(row=1, column=idx, value=col)
        for idx, width in {1: 36, 2: 20, 3: 80, 4: 20, 5: 14, 6: 14}.items():
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.freeze_panes = None
        return

    working = accessorial_df.copy()
    special_mask = working["Cost Name"].astype(str).str.startswith(HEALTHINEERS_SPECIAL_ACCESSORIAL_PREFIXES)
    normal_df = working[~special_mask]
    special_df = working[special_mask]

    current_row = 1
    if not normal_df.empty:
        for idx, col in enumerate(PREON_ACCESSORIAL_COLUMNS, start=1):
            ws.cell(row=current_row, column=idx, value=col)
        data_rows = list(dataframe_to_rows(normal_df, index=False, header=False))
        for row_idx, row in enumerate(data_rows, start=current_row + 1):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        data_end_row = current_row + len(data_rows)
        apply_table_sheet_formatting(
            ws,
            header_row=current_row,
            column_count=len(PREON_ACCESSORIAL_COLUMNS),
            data_end_row=max(current_row, data_end_row),
            column_max_widths={1: 36, 2: 20, 3: 80, 4: 20, 5: 14, 6: 14},
        )
        current_row = data_end_row + 2

    if special_df.empty:
        return

    thin = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=10)
    body_font = Font(size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    key_order: list[tuple[str, str]] = list(
        dict.fromkeys((str(r["Cost Name"]), str(r["Rate by"])) for _, r in special_df.iterrows())
    )
    grouped = special_df.groupby(["Cost Name", "Rate by"], sort=False)

    for cost_name, rate_by in key_order:
        group = grouped.get_group((cost_name, rate_by))
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        ws.cell(current_row, 1, cost_name)
        ws.cell(current_row, 1).font = title_font
        ws.cell(current_row, 1).alignment = center

        ws.merge_cells(start_row=current_row + 1, start_column=1, end_row=current_row + 1, end_column=4)
        ws.cell(current_row + 1, 1, "Applies if invoiced by")
        ws.cell(current_row + 1, 1).font = header_font
        ws.cell(current_row + 1, 1).alignment = center

        ws.merge_cells(start_row=current_row + 2, start_column=1, end_row=current_row + 2, end_column=4)
        ws.cell(current_row + 2, 1, f"Rate by:\n{rate_by}\nRegular rule")
        ws.cell(current_row + 2, 1).font = body_font
        ws.cell(current_row + 2, 1).alignment = left

        headers = ("Lane #", "Currency", "p/unit", "Applies if")
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(current_row + 3, idx, header)
            cell.font = header_font
            cell.alignment = center
            cell.border = thin

        lane_row = current_row + 4
        for lane_idx, (_, row) in enumerate(group.iterrows(), start=1):
            currency, amount = _split_currency_and_amount(_normalize_text(row.get("Cost")))
            values = (
                lane_idx,
                currency,
                amount,
                _normalize_text(row.get("Apply over")),
            )
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(lane_row, col_idx, value)
                cell.font = body_font
                cell.alignment = left if col_idx == 4 else center
                cell.border = thin
            lane_row += 1

        for r in range(current_row, lane_row):
            for c in range(1, 5):
                ws.cell(r, c).border = thin

        current_row = lane_row + 1

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 90
    ws.freeze_panes = None


def _split_currency_and_amount(cost_text: str) -> tuple[str, str]:
    text = _normalize_text(cost_text)
    if not text:
        return "", ""
    parts = text.split(maxsplit=1)
    if len(parts) == 2 and parts[0].isalpha() and len(parts[0]) == 3:
        return parts[0].upper(), parts[1]
    return "", text


def _format_sheet(
    ws,
    shipment_count: int,
    total_cols: int,
    data_start_row: int,
    data_end_row: int,
    group_ranges: list[tuple[int, int]],
    *,
    shipment_columns: list[str],
    bold_headers: set[str],
) -> None:
    thin = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    ship_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    cost_fill = PatternFill(fill_type="solid", fgColor="E2EFDA")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    header_font = Font(bold=True, size=10)
    normal_header = Font(bold=False, size=10)
    body_font = Font(size=10)

    for r in range(1, data_start_row):
        for c in range(1, total_cols + 1):
            cell = ws.cell(r, c)
            cell.border = thin
            cell.alignment = center
            if r == 4 and c <= shipment_count:
                label = shipment_columns[c - 1]
                cell.font = header_font if label in bold_headers else normal_header
                cell.fill = ship_fill
            else:
                cell.font = header_font if r < 4 else normal_header
                cell.fill = ship_fill if c <= shipment_count else cost_fill

    currency_cols = {start for start, _ in group_ranges}
    for r in range(data_start_row, data_end_row + 1):
        for c in range(1, total_cols + 1):
            cell = ws.cell(r, c)
            cell.border = thin
            cell.font = body_font
            if c <= shipment_count:
                cell.alignment = left
            elif c in currency_cols:
                cell.alignment = center
            else:
                cell.alignment = right

    for c in range(1, total_cols + 1):
        letter = get_column_letter(c)
        max_len = 0
        for r in range(1, data_end_row + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v).split("\n")[0]))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 34)

    ws.freeze_panes = ws.cell(row=data_start_row, column=1)
