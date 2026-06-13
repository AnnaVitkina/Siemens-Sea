"""Transform extracted rate data into matrix-format rate cards."""

from __future__ import annotations

import math
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from carrier_lookup import (
    build_all_carriers_apply_if,
    build_ebs_carrier_apply_if,
    get_ebs_carrier_variants,
)
from accessorial_costs import ACCESSORIAL_COLUMNS, build_accessorial_costs_dataframe, write_accessorial_costs_sheet
from cleaning import clean_digi_fcl_rates
from config import (
    BCN_EQUIPMENT_TYPE_VALUE,
    BCN_MULTIPLIER_LABEL,
    BCN_POD_COLUMN,
    FCL_BASE_TAB,
    OUTPUT_DIR,
    RATE_CARD_SHEET_NAME,
)
from extractor import (
    SubfolderSelection,
    extract_sheet_to_dataframe,
    list_excel_files,
    load_processing_context,
    slugify,
)
from glossary_lookup import GlossaryFeeLookup
from rates_surcharge_lookup import RatesSurchargeLookup, load_rates_surcharge_lookup
from thc_lookup import FclThcLookup, load_fcl_thc_lookup

EQUIPMENT_TYPE_VALUE = "not LTL/Buyer Consolidation"

FCL_SHIPMENT_COLUMNS = [
    "Line ID",
    "Carrier SCAC Code",
    "Origin Country",
    "POL",
    "Destination Country",
    "POD",
    "Scope",
    "Service Text",
    "Valid from",
    "Valid to",
    "Equipment type",
]

BCN_SHIPMENT_COLUMNS = [
    "Line ID",
    "Carrier SCAC Code",
    "Origin Country",
    "POL",
    "Destination Country",
    BCN_POD_COLUMN,
    "Scope",
    "Service Text",
    "Valid from",
    "Valid to",
    "Measurement Type",
    "Equipment type",
]

# Backward-compatible alias.
SHIPMENT_COLUMNS = FCL_SHIPMENT_COLUMNS

FCL_BOLD_SHIPMENT_HEADERS = {
    "Carrier SCAC Code",
    "Origin Country",
    "POL",
    "Destination Country",
    "POD",
    "Valid from",
    "Valid to",
    "Equipment type",
}

BCN_BOLD_SHIPMENT_HEADERS = {
    "Carrier SCAC Code",
    "Origin Country",
    "POL",
    "Destination Country",
    BCN_POD_COLUMN,
    "Valid from",
    "Valid to",
    "Measurement Type",
    "Equipment type",
}

CURRENCY_COLUMN_HEADER = "Currency"


@dataclass(frozen=True)
class RateCardFlowProfile:
    flow: str
    shipment_columns: tuple[str, ...]
    bold_shipment_headers: frozenset[str]
    equipment_type_value: str
    cost_header_rows: int
    multiplier_label: str | None
    pod_output_column: str
    apply_all_carriers: bool
    include_measurement_type: bool


def get_rate_card_flow_profile(flow: str) -> RateCardFlowProfile:
    if flow == "BCN":
        return RateCardFlowProfile(
            flow="BCN",
            shipment_columns=tuple(BCN_SHIPMENT_COLUMNS),
            bold_shipment_headers=frozenset(BCN_BOLD_SHIPMENT_HEADERS),
            equipment_type_value=BCN_EQUIPMENT_TYPE_VALUE,
            cost_header_rows=5,
            multiplier_label=BCN_MULTIPLIER_LABEL,
            pod_output_column=BCN_POD_COLUMN,
            apply_all_carriers=True,
            include_measurement_type=True,
        )
    return RateCardFlowProfile(
        flow="FCL",
        shipment_columns=tuple(FCL_SHIPMENT_COLUMNS),
        bold_shipment_headers=frozenset(FCL_BOLD_SHIPMENT_HEADERS),
        equipment_type_value=EQUIPMENT_TYPE_VALUE,
        cost_header_rows=4,
        multiplier_label=None,
        pod_output_column="POD",
        apply_all_carriers=False,
        include_measurement_type=False,
    )


def _split_port_code(port_code: object) -> tuple[str | None, str | None]:
    if pd.isna(port_code):
        return None, None

    code = str(port_code).strip().upper()
    if len(code) >= 5:
        return code[:2], code[-3:]
    if len(code) == 2:
        return code, None
    if len(code) == 3:
        return None, code
    return code[:2] if len(code) >= 2 else None, code[-3:] if len(code) >= 3 else code


def _format_date(value: object) -> str | None:
    if pd.isna(value):
        return None

    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")

    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return str(value)
    return parsed.strftime("%d.%m.%Y")


def _display_container_type(container_type_code: str) -> str:
    return str(container_type_code).replace("_GP", "_FT")


def _cost_column_name(container_type_code: str) -> str:
    return f"Transport cost ({_display_container_type(container_type_code)})"


def _thc_inbound_column_name(container_type_code: str) -> str:
    return f"THC Fee ({container_type_code} FOB INBOUND)"


def _thc_outbound_column_name(container_type_code: str) -> str:
    return f"THC Fee ({container_type_code}, CFR OUTBOUND)"


def _thc_othc_column_name(container_type_code: str) -> str:
    return f"THC Fee ({container_type_code} OTHC)"


def _thc_dthc_column_name(container_type_code: str) -> str:
    return f"THC Fee ({container_type_code} DTHC)"


def _imo_charges_column_name(container_type_code: str) -> str:
    return f"IMO Charges ({_display_container_type(container_type_code)})"


def _ets_fee_column_name(container_type_code: str) -> str:
    return f"ETS Fee ({_display_container_type(container_type_code)})"


def _ebs_carrier_variant_column_name(cost_label: str, container_type_code: str) -> str:
    return f"{cost_label} ({_display_container_type(container_type_code)})"


def _wrs_column_name(container_type_code: str) -> str:
    return f"War Risk Surcharge ({_display_container_type(container_type_code)})"


def _rate_by_value(container_type_code: str) -> str:
    return f"Container/{_display_container_type(container_type_code)}"


def _thc_rate_by_value(container_type_code: str) -> str:
    return f"Container/{container_type_code}"


def _rate_type(calculation_basis: object) -> str:
    if str(calculation_basis).strip().lower() == "container":
        return "p/unit"
    return "Flat"


def _normalize_number(value: object) -> object:
    if pd.isna(value):
        return None
    if isinstance(value, (int, float)):
        if float(value).is_integer():
            return int(value)
        return float(value)
    return value


def _number_format(value: object) -> str:
    if isinstance(value, (int, float)) and float(value).is_integer():
        return "0"
    if isinstance(value, float):
        return "0.##########"
    return "General"


def _container_type_sort_key(container_type_code: str) -> tuple:
    parts = str(container_type_code).split("_")
    size = int(parts[0]) if parts and parts[0].isdigit() else 0
    return size, str(container_type_code)


def _measurement_type(line_id: object) -> str:
    if pd.isna(line_id):
        return "Not Reefer"
    return "Reefer" if str(line_id).strip().upper().startswith("R") else "Not Reefer"


def _prepare_source_dataframe(
    df: pd.DataFrame,
    profile: RateCardFlowProfile,
) -> pd.DataFrame:
    required_columns = [
        "Line ID",
        "Carrier SCAC Code",
        "POL",
        "POD",
        "Scope",
        "Service Text",
        "Valid from",
        "Valid until",
        "Container Type Code",
        "Currency Freight Rate",
        "Cost",
        "Calculation Basis",
    ]
    missing = [column for column in required_columns if column not in df.columns]
    if missing:
        raise KeyError(f"DIGI_FCL_Rates is missing required columns: {', '.join(missing)}")

    prepared = df.copy()
    origin_parts = prepared["POL"].apply(_split_port_code)
    destination_parts = prepared["POD"].apply(_split_port_code)

    prepared["Origin Country"] = origin_parts.apply(lambda part: part[0])
    prepared["POL Code"] = origin_parts.apply(lambda part: part[1])
    prepared["Destination Country"] = destination_parts.apply(lambda part: part[0])
    prepared["POD Code"] = destination_parts.apply(lambda part: part[1])
    prepared["Valid from Fmt"] = prepared["Valid from"].apply(_format_date)
    prepared["Valid to Fmt"] = prepared["Valid until"].apply(_format_date)
    prepared["Equipment type"] = profile.equipment_type_value
    if profile.include_measurement_type:
        prepared["Measurement Type"] = prepared["Line ID"].apply(_measurement_type)

    return prepared


def _build_transport_thc_column_groups(
    prepared: pd.DataFrame,
    container_types: list[str],
    use_othc_dthc_labels: bool = False,
) -> list[dict[str, str]]:
    column_groups: list[dict[str, str]] = []
    for container_type in container_types:
        rows = prepared[prepared["Container Type Code"] == container_type]
        if rows.empty:
            continue
        first_row = rows.iloc[0]
        rate_type = _rate_type(first_row["Calculation Basis"])
        column_groups.extend(
            [
                {
                    "container_type": container_type,
                    "cost_name": _cost_column_name(container_type),
                    "apply_if": "",
                    "rate_by": _rate_by_value(container_type),
                    "rate_type": rate_type,
                },
                {
                    "container_type": container_type,
                    "cost_name": (
                        _thc_othc_column_name(container_type)
                        if use_othc_dthc_labels
                        else _thc_inbound_column_name(container_type)
                    ),
                    "apply_if": "",
                    "rate_by": _thc_rate_by_value(container_type),
                    "rate_type": "p/unit",
                },
                {
                    "container_type": container_type,
                    "cost_name": (
                        _thc_dthc_column_name(container_type)
                        if use_othc_dthc_labels
                        else _thc_outbound_column_name(container_type)
                    ),
                    "apply_if": "",
                    "rate_by": _thc_rate_by_value(container_type),
                    "rate_type": "p/unit",
                },
            ]
        )
    return column_groups


def _build_surcharge_column_groups(
    container_types: list[str],
    shipper: str,
    flow: str,
    profile: RateCardFlowProfile,
) -> list[dict[str, str]]:
    column_groups: list[dict[str, str]] = []
    for container_type in container_types:
        column_groups.append(
            {
                "container_type": container_type,
                "cost_name": _imo_charges_column_name(container_type),
                "apply_if": "",
                "rate_by": _thc_rate_by_value(container_type),
                "rate_type": "p/unit",
            }
        )
    for container_type in container_types:
        column_groups.append(
            {
                "container_type": container_type,
                "cost_name": _ets_fee_column_name(container_type),
                "apply_if": "",
                "rate_by": _thc_rate_by_value(container_type),
                "rate_type": "p/unit",
            }
        )
    for variant in get_ebs_carrier_variants(shipper, flow=flow):
        for container_type in container_types:
            apply_if = ""
            if not profile.apply_all_carriers:
                apply_if = build_ebs_carrier_apply_if(variant["carrier_code"])
            column_groups.append(
                {
                    "container_type": container_type,
                    "cost_name": _ebs_carrier_variant_column_name(
                        variant["cost_label"],
                        container_type,
                    ),
                    "apply_if": apply_if,
                    "rate_by": _thc_rate_by_value(container_type),
                    "rate_type": "p/unit",
                }
            )
    for container_type in container_types:
        column_groups.append(
            {
                "container_type": container_type,
                "cost_name": _wrs_column_name(container_type),
                "apply_if": "",
                "rate_by": _thc_rate_by_value(container_type),
                "rate_type": "p/unit",
            }
        )
    return column_groups


def _lookup_thc_series(
    shipment_df: pd.DataFrame,
    container_type: str,
    currency_series: pd.Series,
    thc_lookup: FclThcLookup,
    direction: str,
) -> pd.Series:
    country_column = "Origin Country" if direction == "inbound" else "Destination Country"
    values: list[object] = []
    for index in range(len(shipment_df)):
        country = shipment_df.iloc[index][country_column]
        currency = currency_series.iloc[index]
        value = thc_lookup.lookup(country, container_type, currency)
        if value is not None:
            value = -abs(float(value))
        values.append(_normalize_number(value))
    return pd.Series(values)


def _lookup_surcharge_columns(
    shipment_df: pd.DataFrame,
    container_type: str,
    surcharge_lookup: RatesSurchargeLookup,
    fee_type: str,
) -> tuple[pd.Series, pd.Series]:
    cost_methods = {
        "imo": surcharge_lookup.lookup_imo,
        "ets": surcharge_lookup.lookup_ets,
        "ebs": surcharge_lookup.lookup_ebs,
        "wrs": surcharge_lookup.lookup_wrs,
    }
    currency_methods = {
        "imo": surcharge_lookup.lookup_imo_currency,
        "ets": surcharge_lookup.lookup_ets_currency,
        "ebs": surcharge_lookup.lookup_ebs_currency,
        "wrs": surcharge_lookup.lookup_wrs_currency,
    }
    lookup_cost = cost_methods[fee_type]
    lookup_currency = currency_methods[fee_type]

    currencies: list[object] = []
    costs: list[object] = []
    for index in range(len(shipment_df)):
        line_id = shipment_df.iloc[index]["Line ID"]
        value = lookup_cost(line_id, container_type)
        if fee_type == "ets" and value is not None:
            value = math.ceil(float(value))
        value = _normalize_number(value)
        costs.append(value)
        currencies.append(lookup_currency(line_id, container_type) if value is not None else None)
    return pd.Series(currencies), pd.Series(costs)


def _optional_pivot_series(
    prepared: pd.DataFrame,
    lane_columns: list[str],
    container_type: str,
    value_column: str,
) -> pd.Series | None:
    if value_column not in prepared.columns:
        return None
    pivot = prepared.pivot_table(
        index=lane_columns,
        columns="Container Type Code",
        values=value_column,
        aggfunc="first",
    )
    if container_type not in pivot.columns:
        return None
    return pivot[container_type].reset_index(drop=True)


def _series_has_values(series: pd.Series) -> bool:
    return bool(series.notna().any())


def _first_matching_column(columns: list[str], patterns: tuple[str, ...]) -> str | None:
    for column in columns:
        upper = column.upper()
        if all(pattern in upper for pattern in patterns):
            return column
    return None


def _ets_source_columns(prepared: pd.DataFrame) -> tuple[str | None, str | None]:
    columns = [str(c) for c in prepared.columns]
    currency_col = _first_matching_column(columns, ("ETS", "CURRENCY"))
    value_col = None
    for column in columns:
        upper = column.upper()
        if "ETS" in upper and "CURRENCY" not in upper:
            value_col = column
            break
    return value_col, currency_col


def _fallback_when_empty(primary: pd.Series, fallback: pd.Series | None) -> pd.Series:
    if _series_has_values(primary) or fallback is None:
        return primary
    return fallback


def _apply_flow_column_group_overrides(
    column_groups: list[dict[str, object]],
    profile: RateCardFlowProfile,
    shipper: str,
) -> None:
    if not profile.apply_all_carriers:
        return

    apply_if = build_all_carriers_apply_if(shipper, flow=profile.flow)
    for group in column_groups:
        group["apply_if"] = apply_if
        if profile.multiplier_label:
            group["multiplier_label"] = profile.multiplier_label


def build_rate_card_dataframe(
    df: pd.DataFrame,
    thc_lookup: FclThcLookup,
    surcharge_lookup: RatesSurchargeLookup,
    shipper: str,
    flow: str,
    glossary_lookups: list[GlossaryFeeLookup] | None = None,
) -> tuple[pd.DataFrame, list[dict[str, object]], RateCardFlowProfile]:
    profile = get_rate_card_flow_profile(flow)
    prepared = _prepare_source_dataframe(df, profile)
    use_othc_dthc_labels = shipper == "Siemens Healthineers LATAM" and flow == "FCL"

    container_types = sorted(
        prepared["Container Type Code"].dropna().astype(str).unique().tolist(),
        key=_container_type_sort_key,
    )
    transport_thc_groups = _build_transport_thc_column_groups(
        prepared,
        container_types,
        use_othc_dthc_labels=use_othc_dthc_labels,
    )
    surcharge_groups = _build_surcharge_column_groups(
        container_types,
        shipper,
        flow,
        profile,
    )
    glossary_groups: list[dict[str, object]] = []
    for glossary_lookup in glossary_lookups or []:
        glossary_groups.extend(glossary_lookup.tmp_fee_column_groups())
        glossary_groups.extend(glossary_lookup.financing_fee_column_groups(container_types))
    column_groups: list[dict[str, object]] = (
        transport_thc_groups + surcharge_groups + glossary_groups
    )
    _apply_flow_column_group_overrides(column_groups, profile, shipper)

    lane_columns = [
        "Line ID",
        "Carrier SCAC Code",
        "Origin Country",
        "POL Code",
        "Destination Country",
        "POD Code",
        "Scope",
        "Service Text",
        "Valid from Fmt",
        "Valid to Fmt",
    ]
    if profile.include_measurement_type:
        lane_columns.append("Measurement Type")
    lane_columns.append("Equipment type")

    cost_pivot = prepared.pivot_table(
        index=lane_columns,
        columns="Container Type Code",
        values="Cost",
        aggfunc="first",
    )
    currency_pivot = prepared.pivot_table(
        index=lane_columns,
        columns="Container Type Code",
        values="Currency Freight Rate",
        aggfunc="first",
    )

    shipment_df = cost_pivot.reset_index()
    rename_map = {
        "POL Code": "POL",
        "POD Code": profile.pod_output_column,
        "Valid from Fmt": "Valid from",
        "Valid to Fmt": "Valid to",
    }
    shipment_df = shipment_df.rename(columns=rename_map)[list(profile.shipment_columns)].copy()

    transport_thc_blocks: list[pd.DataFrame] = []
    imo_blocks: list[pd.DataFrame] = []
    ets_blocks: list[pd.DataFrame] = []
    ebs_variants = get_ebs_carrier_variants(shipper, flow=flow)
    ebs_variant_blocks: dict[str, list[pd.DataFrame]] = {
        variant["key"]: [] for variant in ebs_variants
    }
    wrs_blocks: list[pd.DataFrame] = []
    tmp_blocks: list[pd.DataFrame] = []
    financing_blocks: list[pd.DataFrame] = []
    ets_value_col, ets_currency_col = _ets_source_columns(prepared)
    for container_type in container_types:
        currency_series = currency_pivot[container_type].reset_index(drop=True)
        cost_series = cost_pivot[container_type].reset_index(drop=True).apply(_normalize_number)
        if use_othc_dthc_labels:
            thc_inbound_series = pd.Series([None] * len(shipment_df))
            thc_outbound_series = pd.Series([None] * len(shipment_df))
        else:
            thc_inbound_series = _lookup_thc_series(
                shipment_df,
                container_type,
                currency_series,
                thc_lookup,
                direction="inbound",
            )
            thc_outbound_series = _lookup_thc_series(
                shipment_df,
                container_type,
                currency_series,
                thc_lookup,
                direction="outbound",
            )
        imo_currency, imo_cost = _lookup_surcharge_columns(
            shipment_df,
            container_type,
            surcharge_lookup,
            fee_type="imo",
        )
        ets_currency, ets_cost = _lookup_surcharge_columns(
            shipment_df,
            container_type,
            surcharge_lookup,
            fee_type="ets",
        )
        ebs_currency, ebs_cost = _lookup_surcharge_columns(
            shipment_df,
            container_type,
            surcharge_lookup,
            fee_type="ebs",
        )
        wrs_currency, wrs_cost = _lookup_surcharge_columns(
            shipment_df,
            container_type,
            surcharge_lookup,
            fee_type="wrs",
        )
        thc_in_fallback = _optional_pivot_series(
            prepared,
            lane_columns,
            container_type,
            "THC indication Origin lump sum",
        )
        thc_out_fallback = _optional_pivot_series(
            prepared,
            lane_columns,
            container_type,
            "THC indication Destination lump sum",
        )
        thc_inbound_series = _fallback_when_empty(thc_inbound_series, thc_in_fallback)
        thc_outbound_series = _fallback_when_empty(thc_outbound_series, thc_out_fallback)
        if not _series_has_values(thc_inbound_series):
            thc_inbound_series = pd.Series([None] * len(shipment_df))
        if not _series_has_values(thc_outbound_series):
            thc_outbound_series = pd.Series([None] * len(shipment_df))

        imo_cost_fallback = _optional_pivot_series(
            prepared,
            lane_columns,
            container_type,
            "IMO Charge",
        )
        imo_currency_fallback = _optional_pivot_series(
            prepared,
            lane_columns,
            container_type,
            "Currency IMO Charge",
        )
        imo_cost = _fallback_when_empty(imo_cost, imo_cost_fallback)
        imo_currency = _fallback_when_empty(imo_currency, imo_currency_fallback)

        ets_cost_fallback = (
            _optional_pivot_series(prepared, lane_columns, container_type, ets_value_col)
            if ets_value_col
            else None
        )
        ets_currency_fallback = (
            _optional_pivot_series(prepared, lane_columns, container_type, ets_currency_col)
            if ets_currency_col
            else None
        )
        ets_cost = _fallback_when_empty(ets_cost, ets_cost_fallback)
        ets_currency = _fallback_when_empty(ets_currency, ets_currency_fallback)
        transport_thc_blocks.append(
            pd.DataFrame(
                {
                    f"{container_type}__transport_currency": currency_series,
                    f"{container_type}__transport_cost": cost_series,
                    f"{container_type}__thc_in_currency": currency_series,
                    f"{container_type}__thc_in_cost": thc_inbound_series,
                    f"{container_type}__thc_out_currency": currency_series,
                    f"{container_type}__thc_out_cost": thc_outbound_series,
                }
            )
        )
        imo_blocks.append(
            pd.DataFrame(
                {
                    f"{container_type}__imo_currency": imo_currency,
                    f"{container_type}__imo_cost": imo_cost,
                }
            )
        )
        ets_blocks.append(
            pd.DataFrame(
                {
                    f"{container_type}__ets_currency": ets_currency,
                    f"{container_type}__ets_cost": ets_cost,
                }
            )
        )

        for variant in ebs_variants:
            variant_key = variant["key"]
            ebs_variant_blocks[variant_key].append(
                pd.DataFrame(
                    {
                        f"{container_type}__ebs_{variant_key}_currency": ebs_currency,
                        f"{container_type}__ebs_{variant_key}_cost": ebs_cost,
                    }
                )
            )
        wrs_blocks.append(
            pd.DataFrame(
                {
                    f"{container_type}__wrs_currency": wrs_currency,
                    f"{container_type}__wrs_cost": wrs_cost,
                }
            )
        )

    transport_present_by_type = {
        container_type: cost_pivot[container_type].reset_index(drop=True).notna()
        for container_type in container_types
    }

    for glossary_lookup in glossary_lookups or []:
        carrier_key = glossary_lookup.carrier_key
        if glossary_lookup.fees.financing_fee is not None:
            for container_type in container_types:
                financing_rate = glossary_lookup.financing_fee_rate(container_type)
                financing_currency = glossary_lookup.financing_fee_currency()
                transport_present = transport_present_by_type[container_type]
                financing_cost_series = pd.Series(
                    [_normalize_number(financing_rate)] * len(shipment_df)
                ).where(transport_present, None)
                financing_currency_series = pd.Series(
                    [financing_currency] * len(shipment_df)
                ).where(transport_present, None)
                financing_blocks.append(
                    pd.DataFrame(
                        {
                            f"{carrier_key}__{container_type}__financing_currency": (
                                financing_currency_series
                            ),
                            f"{carrier_key}__{container_type}__financing_cost": financing_cost_series,
                        }
                    )
                )

        for index, entry in enumerate(glossary_lookup.fees.tmp_fees):
            row_count = len(shipment_df)
            tmp_blocks.append(
                pd.DataFrame(
                    {
                        f"tmp_{carrier_key}_{index}_currency": [entry.currency] * row_count,
                        f"tmp_{carrier_key}_{index}_unit": [_normalize_number(entry.unit_rate)]
                        * row_count,
                        f"tmp_{carrier_key}_{index}_flat": [_normalize_number(entry.max_cap)]
                        * row_count,
                    }
                )
            )

    ebs_blocks_ordered = [
        block
        for variant in ebs_variants
        for block in ebs_variant_blocks[variant["key"]]
    ]
    rate_card_df = pd.concat(
        [
            shipment_df,
            *transport_thc_blocks,
            *imo_blocks,
            *ets_blocks,
            *ebs_blocks_ordered,
            *wrs_blocks,
            *tmp_blocks,
            *financing_blocks,
        ],
        axis=1,
    )
    return rate_card_df, column_groups, profile


def build_fcl_rate_card_dataframe(
    df: pd.DataFrame,
    thc_lookup: FclThcLookup,
    surcharge_lookup: RatesSurchargeLookup,
    shipper: str,
    glossary_lookups: list[GlossaryFeeLookup] | None = None,
) -> tuple[pd.DataFrame, list[dict[str, object]]]:
    rate_card_df, column_groups, _profile = build_rate_card_dataframe(
        df,
        thc_lookup,
        surcharge_lookup,
        shipper,
        "FCL",
        glossary_lookups=glossary_lookups,
    )
    return rate_card_df, column_groups


def _column_group_width(meta: dict[str, object]) -> int:
    return int(meta.get("column_count", 2))


def _apply_rate_card_formatting(
    worksheet,
    shipment_count: int,
    total_columns: int,
    data_start_row: int,
    data_end_row: int,
    cost_column_groups: list[tuple[int, ...]],
    profile: RateCardFlowProfile,
) -> None:
    thin_border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    shipment_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    cost_fill = PatternFill(fill_type="solid", fgColor="E2EFDA")
    header_font = Font(bold=True, size=10)
    normal_header_font = Font(bold=False, size=10)
    body_font = Font(size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    number_alignment = Alignment(horizontal="right", vertical="center")

    for row_index in range(1, data_start_row):
        for column_index in range(1, total_columns + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.border = thin_border
            cell.alignment = center
            if row_index == profile.cost_header_rows and column_index <= shipment_count:
                header_name = profile.shipment_columns[column_index - 1]
                cell.font = (
                    header_font
                    if header_name in profile.bold_shipment_headers
                    else normal_header_font
                )
                cell.fill = shipment_fill
            else:
                cell.font = (
                    header_font if row_index < profile.cost_header_rows else normal_header_font
                )
                cell.fill = cost_fill if column_index > shipment_count else shipment_fill

    for row_index in range(data_start_row, data_end_row + 1):
        for column_index in range(1, total_columns + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.border = thin_border
            cell.font = body_font
            if column_index <= shipment_count:
                cell.alignment = left
            elif _is_currency_column(column_index, shipment_count, cost_column_groups):
                cell.alignment = center
            else:
                cell.alignment = number_alignment
                if isinstance(cell.value, (int, float)):
                    cell.number_format = _number_format(cell.value)

    for column_index in range(1, total_columns + 1):
        column_letter = get_column_letter(column_index)
        max_length = 0
        for row_index in range(1, data_end_row + 1):
            value = worksheet.cell(row=row_index, column=column_index).value
            if value is None:
                continue
            max_length = max(max_length, len(str(value).split("\n")[0]))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 28)

    worksheet.freeze_panes = worksheet.cell(row=data_start_row, column=1)
    worksheet.auto_filter.ref = None


def _is_currency_column(
    column_index: int,
    shipment_count: int,
    cost_column_groups: list[tuple[int, ...]],
) -> bool:
    for group in cost_column_groups:
        if column_index == group[0]:
            return True
    return False


def _write_rate_card_sheet(
    workbook: Workbook,
    rate_card_df: pd.DataFrame,
    column_groups: list[dict[str, object]],
    profile: RateCardFlowProfile,
) -> None:
    worksheet = workbook.active
    worksheet.title = RATE_CARD_SHEET_NAME

    shipment_count = len(profile.shipment_columns)
    cost_column_groups: list[tuple[int, ...]] = []
    next_column = shipment_count + 1
    rate_by_row = 4 if profile.multiplier_label else 3
    currency_row = profile.cost_header_rows

    for meta in column_groups:
        column_count = _column_group_width(meta)
        group_columns = tuple(range(next_column, next_column + column_count))
        cost_column_groups.append(group_columns)
        currency_col = group_columns[0]

        for row_index in (1, 2):
            worksheet.merge_cells(
                start_row=row_index,
                start_column=currency_col,
                end_row=row_index,
                end_column=group_columns[-1],
            )

        if profile.multiplier_label:
            worksheet.merge_cells(
                start_row=3,
                start_column=currency_col,
                end_row=3,
                end_column=group_columns[-1],
            )
            worksheet.cell(
                row=3,
                column=currency_col,
                value=meta.get("multiplier_label", profile.multiplier_label),
            )

        row3_flat_label = meta.get("row3_flat_label")
        if column_count == 3 and row3_flat_label:
            worksheet.merge_cells(
                start_row=rate_by_row,
                start_column=currency_col,
                end_row=rate_by_row,
                end_column=currency_col + 1,
            )
            worksheet.cell(row=rate_by_row, column=currency_col, value=meta["rate_by"])
            worksheet.cell(row=rate_by_row, column=group_columns[-1], value=row3_flat_label)
        else:
            worksheet.merge_cells(
                start_row=rate_by_row,
                start_column=currency_col,
                end_row=rate_by_row,
                end_column=group_columns[-1],
            )
            worksheet.cell(row=rate_by_row, column=currency_col, value=meta["rate_by"])

        worksheet.cell(row=1, column=currency_col, value=meta["cost_name"])
        worksheet.cell(row=2, column=currency_col, value=meta["apply_if"])
        worksheet.cell(row=currency_row, column=currency_col, value=CURRENCY_COLUMN_HEADER)

        rate_types = meta.get("rate_types")
        if rate_types:
            for offset, rate_type in enumerate(rate_types, start=1):
                worksheet.cell(
                    row=currency_row,
                    column=currency_col + offset,
                    value=rate_type,
                )
        else:
            worksheet.cell(
                row=currency_row,
                column=currency_col + 1,
                value=meta["rate_type"],
            )

        next_column += column_count

    for column_index, column_name in enumerate(profile.shipment_columns, start=1):
        worksheet.cell(row=currency_row, column=column_index, value=column_name)

    data_start_row = profile.cost_header_rows + 1
    data_rows = list(dataframe_to_rows(rate_card_df, index=False, header=False))
    for row_offset, row in enumerate(data_rows, start=data_start_row):
        for column_offset, value in enumerate(row, start=1):
            worksheet.cell(row=row_offset, column=column_offset, value=value)

    total_columns = shipment_count + sum(_column_group_width(meta) for meta in column_groups)
    data_end_row = data_start_row + len(data_rows) - 1
    _apply_rate_card_formatting(
        worksheet=worksheet,
        shipment_count=shipment_count,
        total_columns=total_columns,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
        cost_column_groups=cost_column_groups,
        profile=profile,
    )


def _prune_empty_cost_groups(
    rate_card_df: pd.DataFrame,
    column_groups: list[dict[str, object]],
    shipment_count: int,
) -> tuple[pd.DataFrame, list[dict[str, object]]]:
    if not column_groups:
        return rate_card_df, column_groups

    kept_groups: list[dict[str, object]] = []
    kept_cost_columns: list[str] = []
    next_index = shipment_count

    for meta in column_groups:
        width = _column_group_width(meta)
        group_cols = list(rate_card_df.columns[next_index : next_index + width])
        next_index += width
        if not group_cols:
            continue

        value_cols = group_cols[1:] if len(group_cols) > 1 else group_cols
        group_frame = rate_card_df[value_cols]
        non_empty = group_frame.apply(
            lambda col: col.notna() & (col.astype(str).str.strip() != "")
        ).any().any()
        if non_empty:
            kept_groups.append(meta)
            kept_cost_columns.extend(group_cols)

    kept_columns = list(rate_card_df.columns[:shipment_count]) + kept_cost_columns
    return rate_card_df[kept_columns].copy(), kept_groups


def build_output_rate_card_path(flow: str, shipper: str) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    shipper_slug = slugify(shipper)
    return OUTPUT_DIR / f"{flow}_{shipper_slug}_rate_card.xlsx"


def save_rate_card(
    df: pd.DataFrame,
    shipper: str,
    flow: str,
    thc_lookup: FclThcLookup | None = None,
    surcharge_lookup: RatesSurchargeLookup | None = None,
    glossary_lookups: list[GlossaryFeeLookup] | None = None,
    individual_selections: list[SubfolderSelection] | None = None,
    output_path: Path | None = None,
) -> tuple[Path, pd.DataFrame]:
    if thc_lookup is None:
        thc_lookup = load_fcl_thc_lookup()
    if surcharge_lookup is None:
        surcharge_lookup = load_rates_surcharge_lookup()

    rate_card_df, column_groups, profile = build_rate_card_dataframe(
        df,
        thc_lookup,
        surcharge_lookup,
        shipper,
        flow,
        glossary_lookups=glossary_lookups,
    )
    rate_card_df, column_groups = _prune_empty_cost_groups(
        rate_card_df,
        column_groups,
        shipment_count=len(profile.shipment_columns),
    )

    if output_path is None:
        output_path = build_output_rate_card_path(flow, shipper)

    accessorial_df = build_accessorial_costs_dataframe(
        shipper,
        individual_selections or [],
        flow=flow,
    )

    workbook = Workbook()
    _write_rate_card_sheet(workbook, rate_card_df, column_groups, profile)
    write_accessorial_costs_sheet(workbook, accessorial_df)
    workbook.save(output_path)

    return output_path, rate_card_df


def save_fcl_rate_card(
    df: pd.DataFrame,
    shipper: str,
    thc_lookup: FclThcLookup | None = None,
    surcharge_lookup: RatesSurchargeLookup | None = None,
    glossary_lookups: list[GlossaryFeeLookup] | None = None,
    individual_selections: list[SubfolderSelection] | None = None,
    output_path: Path | None = None,
) -> tuple[Path, pd.DataFrame]:
    return save_rate_card(
        df,
        shipper,
        "FCL",
        thc_lookup=thc_lookup,
        surcharge_lookup=surcharge_lookup,
        glossary_lookups=glossary_lookups,
        individual_selections=individual_selections,
        output_path=output_path,
    )


def load_digi_fcl_rates_dataframe(
    processing_path: Path | None = None,
    source_file: Path | None = None,
) -> pd.DataFrame:
    if processing_path and processing_path.exists():
        return pd.read_excel(processing_path, sheet_name=FCL_BASE_TAB)

    if source_file and source_file.exists():
        extracted = extract_sheet_to_dataframe(source_file, FCL_BASE_TAB)
        return clean_digi_fcl_rates(extracted)

    context = load_processing_context()
    if context:
        context_path = Path(context["output_path"])
        if context_path.exists():
            return pd.read_excel(context_path, sheet_name=FCL_BASE_TAB)

    main_rate_files = list_excel_files("main rates")
    for file_path in main_rate_files:
        sheet_names = pd.ExcelFile(file_path).sheet_names
        if FCL_BASE_TAB in sheet_names:
            extracted = extract_sheet_to_dataframe(file_path, FCL_BASE_TAB)
            return clean_digi_fcl_rates(extracted)

    raise FileNotFoundError(
        "Could not locate DIGI_FCL_Rates data. Run main.py extraction first "
        "or place a main rates file in the input folder."
    )


def resolve_shipper(shipper: str | None = None) -> str:
    if shipper:
        return shipper

    context = load_processing_context()
    if context and context.get("shipper"):
        return context["shipper"]

    return "unknown"
