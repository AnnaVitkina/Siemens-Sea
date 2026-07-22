"""Build LCL rate cards from individual-rate LCL_Rates tabs."""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from accessorial_costs import (
    apply_table_sheet_formatting,
    build_lcl_accessorial_costs_dataframe,
    write_accessorial_costs_sheet,
)
from carrier_lookup import carrier_code_from_filename, detect_carrier_key
from date_utils import format_dd_mm_yyyy
from config import (
    CONDITIONAL_RULES_SHEET_NAME,
    GLOSSARY_TAB,
    INDIVIDUAL_RATE_SUBFOLDER,
    OUTPUT_DIR,
    RATE_CARD_SHEET_NAME,
)
from extractor import (
    SubfolderSelection,
    extract_sheet_to_dataframe,
    load_processing_context,
    read_workbook_sheet_names,
    slugify,
)

LCL_COST_HEADER_ROWS = 4
CURRENCY_COLUMN_HEADER = "Currency"

DEDUPE_KEY_COLUMNS = (
    "Origin Country",
    "Origin CFS Code",
    "Destination Country",
    "Destination CFS Code",
    "Valid From",
    "Valid to",
)

BOLD_SHIPMENT_HEADERS = {
    "Origin Country",
    "Origin CFS Code",
    "Origin CFS code",
    "Destination Country",
    "Destination CFS Code",
    "Destination CFS code",
    "Valid From",
    "Valid to",
}

COST_TYPE_ALIASES = {
    "LCL RATE": "transport",
    "EU ETS SURCHARGE": "ets",
    "IMO CHARGE": "imo",
    "RED SEA SURCHARGE": "red_sea",
    "LCL RATE ADDER 100% CO2 INSETTING": "biofuel",
}

LCL_COST_GROUPS = (
    {
        "key": "transport",
        "cost_name": "Transport cost (LCL Rate)",
        "apply_if": "LTL/Standard",
        "rate_by": "Volume/cbm",
        "include_flat_min": True,
    },
    {
        "key": "ets",
        "cost_name": "ETS Fee",
        "apply_if": "LTL/Standard",
        "rate_by": "Volume/cbm",
        "include_flat_min": False,
    },
    {
        "key": "imo",
        "cost_name": "IMO Surcharge",
        "apply_if": "LTL/Standard",
        "rate_by": "Volume/cbm",
        "include_flat_min": True,
    },
    {
        "key": "red_sea",
        "cost_name": "Red Sea Surcharge",
        "apply_if": "LTL/Standard",
        "rate_by": "Volume/cbm",
        "include_flat_min": False,
    },
    {
        "key": "biofuel",
        "cost_name": "Biofuel Add-on",
        "apply_if": "LTL/Standard",
        "rate_by": "Volume/cbm",
        "include_flat_min": True,
    },
)

CONDITIONAL_RULES_COLUMNS = ("Column name", "Value", "Conditional Rule")
ORIGIN_CFS_COLUMN = "Origin CFS Code"
DESTINATION_CFS_COLUMN = "Destination CFS Code"
ALL_PORTS_VALUE = "ALL PORTS"
ALPHABET_RULE = ", ".join(f"'{letter}'" for letter in "ABCDEFGHIJKLMNOPQRSTUVWXYZ")


@dataclass(frozen=True)
class ParsedLclRow:
    tab_name: str
    shipment_values: dict[str, object]
    cost_kind: str
    currency: str | None
    cost_value: object


def is_lcl_rates_tab(tab_name: str) -> bool:
    normalized = re.sub(r"[\s_]+", "", tab_name.lower())
    return "lclrate" in normalized or "ltlrate" in normalized


def resolve_lcl_rates_tabs(file_path: Path) -> list[str]:
    return [tab for tab in read_workbook_sheet_names(file_path) if is_lcl_rates_tab(tab)]


def resolve_lcl_tabs(file_path: Path) -> list[str]:
    tabs = resolve_lcl_rates_tabs(file_path)
    if GLOSSARY_TAB in read_workbook_sheet_names(file_path):
        tabs.append(GLOSSARY_TAB)
    return tabs


def normalize_column_name(name: object) -> str:
    return re.sub(r"\s+", " ", str(name).replace("\n", " ")).strip()


def _detect_lcl_rates_header_row(raw_df: pd.DataFrame) -> int:
    for idx in range(min(30, len(raw_df))):
        row_values = {
            normalize_column_name(value).lower()
            for value in raw_df.iloc[idx].tolist()
            if pd.notna(value) and str(value).strip()
        }
        if "cost type" in row_values and (
            "origin cfs code" in row_values
            or "lane id" in row_values
            or "origin country" in row_values
        ):
            return idx
    return 0


def _read_excel_engine(file_path: Path) -> str | None:
    if file_path.suffix.lower() == ".xlsb":
        return "pyxlsb"
    return None


def read_lcl_rates_tab_dataframe(file_path: Path, tab_name: str) -> pd.DataFrame:
    engine = _read_excel_engine(file_path)
    raw = pd.read_excel(file_path, sheet_name=tab_name, header=None, engine=engine)
    header_row = _detect_lcl_rates_header_row(raw)
    df = pd.read_excel(file_path, sheet_name=tab_name, header=header_row, engine=engine)
    df = normalize_dataframe_columns(df)
    return df.dropna(axis=1, how="all").dropna(axis=0, how="all")


def _find_cost_type_column(df: pd.DataFrame) -> str | None:
    for column in df.columns:
        if normalize_column_name(column).lower() == "cost type":
            return column
    return None


def normalize_dataframe_columns(df: pd.DataFrame) -> pd.DataFrame:
    renamed = {column: normalize_column_name(column) for column in df.columns}
    return df.rename(columns=renamed)


def _format_lcl_date(value: object) -> str | None:
    return format_dd_mm_yyyy(value)


def _normalize_number(value: object) -> object:
    if pd.isna(value):
        return None
    if isinstance(value, (int, float)):
        if float(value).is_integer():
            return int(value)
        return float(value)
    return value


def _number_format(value: object) -> str:
    if isinstance(value, (int, float)) and float(value).is_integer():
        return "0"
    if isinstance(value, float):
        return "0.##########"
    return "General"


def _columns_before_cost_type(df: pd.DataFrame) -> list[str]:
    columns: list[str] = []
    for column in df.columns:
        if normalize_column_name(column).lower() == "cost type":
            break
        columns.append(column)
    return columns


def _find_date_source_columns(df: pd.DataFrame) -> tuple[str | None, str | None]:
    valid_from_column = None
    valid_until_column = None
    for column in df.columns:
        normalized = normalize_column_name(column).lower()
        if normalized == "valid from":
            valid_from_column = column
        elif normalized == "valid until":
            valid_until_column = column
    return valid_from_column, valid_until_column


def _find_cost_column(df: pd.DataFrame) -> str | None:
    for column in df.columns:
        normalized = normalize_column_name(column).lower()
        if normalized in {"cost", "cost actual rate", "green cost"}:
            return column
    columns = list(df.columns)
    if "Currency LCL" in columns:
        index = columns.index("Currency LCL")
        if index + 1 < len(columns):
            return columns[index + 1]
    return None


def _find_optional_surcharge_column(df: pd.DataFrame, patterns: tuple[str, ...]) -> str | None:
    for column in df.columns:
        normalized = normalize_column_name(column).lower()
        if any(pattern in normalized for pattern in patterns):
            return column
    return None


def _normalize_cost_type(value: object) -> str | None:
    if pd.isna(value):
        return None
    normalized = re.sub(r"\s+", " ", str(value).strip()).upper()
    return COST_TYPE_ALIASES.get(normalized)


def _shipment_column_order(
    pre_cost_columns: list[str],
    *,
    has_valid_from: bool = False,
    has_valid_until: bool = False,
) -> list[str]:
    ordered: list[str] = ["Tab Name"]
    seen = {"Tab Name"}

    preferred = [
        "Source",
        "Service Code",
        "Mode",
        "Supplier Name",
        "Line ID",
        "Lane ID",
        "Item ID",
        "ITEM ID",
        "Origin Country",
        "Origin Region",
        "Origin CFS Code",
        "Origin CFS Name",
        "Destination Country",
        "Destination Region",
        "Destination CFS Code",
        "Destination CFS Name",
    ]
    normalized_lookup = {normalize_column_name(column): column for column in pre_cost_columns}

    for preferred_name in preferred:
        actual = normalized_lookup.get(preferred_name)
        if actual and actual not in seen:
            ordered.append(actual)
            seen.add(actual)

    for column in pre_cost_columns:
        if column not in seen and normalize_column_name(column) not in {"Valid from", "Valid until"}:
            ordered.append(column)
            seen.add(column)

    if has_valid_from and "Valid From" not in seen:
        ordered.append("Valid From")
        seen.add("Valid From")
    if has_valid_until and "Valid to" not in seen:
        ordered.append("Valid to")
        seen.add("Valid to")

    return ordered


def _build_shipment_values(
    row: pd.Series,
    tab_name: str,
    pre_cost_columns: list[str],
    shipment_column_order: list[str],
    valid_from_column: str | None,
    valid_until_column: str | None,
) -> dict[str, object]:
    values: dict[str, object] = {"Tab Name": tab_name}
    for column in pre_cost_columns:
        values[column] = row[column]

    if valid_from_column is not None:
        values["Valid From"] = _format_lcl_date(row[valid_from_column])
    if valid_until_column is not None:
        values["Valid to"] = _format_lcl_date(row[valid_until_column])

    return {column: values.get(column) for column in shipment_column_order}


def _dedupe_key(shipment_values: dict[str, object]) -> tuple:
    return tuple(shipment_values.get(column) for column in DEDUPE_KEY_COLUMNS)


def _resolve_cost_value(
    row: pd.Series,
    df: pd.DataFrame,
    cost_kind: str,
    cost_column: str | None,
) -> object:
    if cost_column is not None:
        primary = _normalize_number(row.get(cost_column))
        if primary is not None:
            return primary

    if cost_kind == "ets":
        ets_column = _find_optional_surcharge_column(df, ("ets",))
        if ets_column is not None:
            return _normalize_number(row.get(ets_column))
    if cost_kind == "red_sea":
        red_sea_column = _find_optional_surcharge_column(df, ("red sea",))
        if red_sea_column is not None:
            return _normalize_number(row.get(red_sea_column))
    return None


def _parse_lcl_rows_from_tab(tab_name: str, df: pd.DataFrame) -> tuple[list[ParsedLclRow], list[str]]:
    cost_type_column = _find_cost_type_column(df)
    if cost_type_column is None:
        return [], ["Tab Name"]

    pre_cost_columns = _columns_before_cost_type(df)
    valid_from_column, valid_until_column = _find_date_source_columns(df)
    shipment_column_order = _shipment_column_order(
        pre_cost_columns,
        has_valid_from=valid_from_column is not None,
        has_valid_until=valid_until_column is not None,
    )
    cost_column = _find_cost_column(df)
    parsed_rows: list[ParsedLclRow] = []
    surcharge_by_key: dict[tuple, dict[str, object]] = {}

    for _, row in df.iterrows():
        cost_kind = _normalize_cost_type(row.get(cost_type_column))
        if cost_kind != "transport":
            continue

        shipment_values = _build_shipment_values(
            row,
            tab_name,
            pre_cost_columns,
            shipment_column_order,
            valid_from_column,
            valid_until_column,
        )
        key = _dedupe_key(shipment_values)
        surcharge_by_key[key] = {
            "ets": _resolve_cost_value(row, df, "ets", cost_column),
            "red_sea": _resolve_cost_value(row, df, "red_sea", cost_column),
        }

    for _, row in df.iterrows():
        cost_kind = _normalize_cost_type(row.get(cost_type_column))
        if cost_kind is None:
            continue

        shipment_values = _build_shipment_values(
            row,
            tab_name,
            pre_cost_columns,
            shipment_column_order,
            valid_from_column,
            valid_until_column,
        )
        currency = row.get("Currency LCL")
        currency_value = None if pd.isna(currency) else str(currency).strip().upper()
        cost_value = _resolve_cost_value(row, df, cost_kind, cost_column)
        if cost_value is None and cost_kind in {"ets", "red_sea"}:
            cost_value = surcharge_by_key.get(_dedupe_key(shipment_values), {}).get(cost_kind)

        parsed_rows.append(
            ParsedLclRow(
                tab_name=tab_name,
                shipment_values=shipment_values,
                cost_kind=cost_kind,
                currency=currency_value,
                cost_value=cost_value,
            )
        )

    return parsed_rows, shipment_column_order


def load_lcl_parsed_rows(selections: list[SubfolderSelection]) -> tuple[list[ParsedLclRow], list[str]]:
    parsed_rows: list[ParsedLclRow] = []
    shipment_columns: list[str] | None = None

    for selection in selections:
        if selection.subfolder != INDIVIDUAL_RATE_SUBFOLDER:
            continue
        for tab in selection.tabs:
            if not is_lcl_rates_tab(tab):
                continue
            df = read_lcl_rates_tab_dataframe(selection.file_path, tab)
            tab_rows, tab_shipment_columns = _parse_lcl_rows_from_tab(tab, df)
            parsed_rows.extend(tab_rows)
            if shipment_columns is None:
                shipment_columns = tab_shipment_columns
            else:
                for column in tab_shipment_columns:
                    if column not in shipment_columns:
                        shipment_columns.append(column)

    return parsed_rows, shipment_columns or ["Tab Name"]


def load_lcl_parsed_rows_from_processing(
    processing_path: Path,
    selections: list[SubfolderSelection],
) -> tuple[list[ParsedLclRow], list[str]]:
    parsed_rows: list[ParsedLclRow] = []
    shipment_columns: list[str] | None = None
    workbook = pd.ExcelFile(processing_path)

    for selection in selections:
        if selection.subfolder != INDIVIDUAL_RATE_SUBFOLDER:
            continue
        for tab in selection.tabs:
            if not is_lcl_rates_tab(tab):
                continue
            sheet_name = _find_processing_sheet_name(workbook.sheet_names, tab)
            if sheet_name is None:
                continue
            raw = pd.read_excel(processing_path, sheet_name=sheet_name, header=None)
            header_row = _detect_lcl_rates_header_row(raw)
            df = normalize_dataframe_columns(
                pd.read_excel(processing_path, sheet_name=sheet_name, header=header_row)
            )
            tab_rows, tab_shipment_columns = _parse_lcl_rows_from_tab(tab, df)
            parsed_rows.extend(tab_rows)
            if shipment_columns is None:
                shipment_columns = tab_shipment_columns
            else:
                for column in tab_shipment_columns:
                    if column not in shipment_columns:
                        shipment_columns.append(column)

    return parsed_rows, shipment_columns or ["Tab Name"]


def _find_processing_sheet_name(sheet_names: list[str], source_tab: str) -> str | None:
    if source_tab in sheet_names:
        return source_tab
    normalized_source = normalize_column_name(source_tab).lower()
    for sheet_name in sheet_names:
        if normalize_column_name(sheet_name).lower().endswith(normalized_source):
            return sheet_name
        if normalized_source in normalize_column_name(sheet_name).lower():
            return sheet_name
    return None


def build_lcl_rate_card_dataframe(
    parsed_rows: list[ParsedLclRow],
    shipment_columns: list[str],
) -> tuple[pd.DataFrame, list[dict[str, object]], list[str]]:
    grouped: dict[tuple, dict[str, object]] = {}
    tab_names_by_key: dict[tuple, set[str]] = {}
    costs_by_key: dict[tuple, dict[str, dict[str, object]]] = {}
    shipment_keys: list[tuple] = []

    for parsed_row in parsed_rows:
        key = _dedupe_key(parsed_row.shipment_values)
        if key not in grouped:
            grouped[key] = dict(parsed_row.shipment_values)
            tab_names_by_key[key] = {parsed_row.tab_name}
            costs_by_key[key] = {}
            shipment_keys.append(key)
        else:
            tab_names_by_key[key].add(parsed_row.tab_name)
            for column, value in parsed_row.shipment_values.items():
                if column == "Tab Name":
                    continue
                if pd.isna(grouped[key].get(column)) and not pd.isna(value):
                    grouped[key][column] = value

        if parsed_row.cost_value is None and parsed_row.currency is None:
            continue

        costs_by_key[key].setdefault(parsed_row.cost_kind, {})
        if parsed_row.currency is not None:
            costs_by_key[key][parsed_row.cost_kind]["currency"] = parsed_row.currency
        if parsed_row.cost_value is not None:
            costs_by_key[key][parsed_row.cost_kind]["cost"] = parsed_row.cost_value

    shipment_rows: list[dict[str, object]] = []
    for key in shipment_keys:
        shipment_values = grouped[key]
        row = {column: shipment_values.get(column) for column in shipment_columns}
        row["Tab Name"] = "; ".join(sorted(tab_names_by_key[key]))
        shipment_rows.append(row)

    shipment_df = pd.DataFrame(shipment_rows, columns=shipment_columns)
    cost_blocks: list[pd.DataFrame] = []
    column_groups = [_build_cost_group_meta(group) for group in LCL_COST_GROUPS]

    for group in LCL_COST_GROUPS:
        currencies: list[object] = []
        flat_mins: list[object] = []
        p_units: list[object] = []
        for key in shipment_keys:
            costs = costs_by_key.get(key, {}).get(group["key"], {})
            currencies.append(costs.get("currency"))
            cost_value = costs.get("cost")
            if group["include_flat_min"]:
                flat_mins.append(cost_value)
            p_units.append(cost_value)

        block_columns: dict[str, object] = {f"{group['key']}__currency": currencies}
        if group["include_flat_min"]:
            block_columns[f"{group['key']}__flat_min"] = flat_mins
        block_columns[f"{group['key']}__p_unit"] = p_units
        cost_blocks.append(pd.DataFrame(block_columns))

    rate_card_df = pd.concat([shipment_df, *cost_blocks], axis=1)
    return rate_card_df, column_groups, shipment_columns


def _build_cost_group_meta(group: dict[str, object]) -> dict[str, object]:
    if group["include_flat_min"]:
        return {
            "cost_name": group["cost_name"],
            "apply_if": group["apply_if"],
            "rate_by": group["rate_by"],
            "column_count": 3,
            "rate_types": ["Flat MIN", "p/unit"],
        }
    return {
        "cost_name": group["cost_name"],
        "apply_if": group["apply_if"],
        "rate_by": group["rate_by"],
        "column_count": 2,
        "rate_type": "p/unit",
    }


def _last_three_letters(value: object) -> str:
    text = str(value).strip().upper()
    return text[-3:] if len(text) >= 3 else text


def _build_all_de_rule(port_var: str, country_code: str, values: set[str]) -> str:
    suffixes = sorted(
        {
            _last_three_letters(item)
            for item in values
            if str(item).strip().upper()[:2] == country_code
        }
    )
    if not suffixes:
        return f"{port_var} equals '{country_code}' in any item"
    equals_part = ", ".join(f"'{suffix}'" for suffix in suffixes)
    return f"{port_var} equals {equals_part} in any item"


def build_conditional_cures_dataframe(
    parsed_rows: list[ParsedLclRow],
) -> pd.DataFrame:
    origin_values = {
        str(row.shipment_values.get(ORIGIN_CFS_COLUMN)).strip()
        for row in parsed_rows
        if not pd.isna(row.shipment_values.get(ORIGIN_CFS_COLUMN))
    }
    destination_values = {
        str(row.shipment_values.get(DESTINATION_CFS_COLUMN)).strip()
        for row in parsed_rows
        if not pd.isna(row.shipment_values.get(DESTINATION_CFS_COLUMN))
    }

    rows: list[dict[str, str]] = []

    for value in sorted(origin_values):
        rule = _build_conditional_rule(ORIGIN_CFS_COLUMN, value, origin_values, destination_values)
        rows.append({"Column name": ORIGIN_CFS_COLUMN, "Value": value, "Conditional Rule": rule})

    for value in sorted(destination_values):
        rule = _build_conditional_rule(
            DESTINATION_CFS_COLUMN,
            value,
            origin_values,
            destination_values,
        )
        rows.append(
            {
                "Column name": DESTINATION_CFS_COLUMN,
                "Value": value,
                "Conditional Rule": rule,
            }
        )

    if not rows:
        return pd.DataFrame(columns=CONDITIONAL_RULES_COLUMNS)

    return pd.DataFrame(rows, columns=CONDITIONAL_RULES_COLUMNS)


def _build_conditional_rule(
    column_name: str,
    value: str,
    origin_values: set[str],
    destination_values: set[str],
) -> str:
    is_origin = column_name == ORIGIN_CFS_COLUMN
    port_var = "SHIP_PORT" if is_origin else "CUST_POST"
    normalized = value.strip().upper()
    source_values = origin_values if is_origin else destination_values

    if normalized == ALL_PORTS_VALUE:
        return f"{port_var} starts with {ALPHABET_RULE} in all items"

    if normalized.startswith("ALL-"):
        country_code = normalized[4:6]
        return _build_all_de_rule(port_var, country_code, source_values)

    last_three = _last_three_letters(value)
    return f"{port_var} equals '{last_three}' in any item"


def _column_group_width(meta: dict[str, object]) -> int:
    return int(meta.get("column_count", 2))


def _is_currency_column(
    column_index: int,
    shipment_count: int,
    cost_column_groups: list[tuple[int, ...]],
) -> bool:
    for group in cost_column_groups:
        if column_index == group[0]:
            return True
    return False


def _apply_rate_card_formatting(
    worksheet,
    shipment_columns: list[str],
    total_columns: int,
    data_start_row: int,
    data_end_row: int,
    cost_column_groups: list[tuple[int, ...]],
) -> None:
    shipment_count = len(shipment_columns)
    thin_border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    shipment_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    cost_fill = PatternFill(fill_type="solid", fgColor="E2EFDA")
    header_font = Font(bold=True, size=10)
    normal_header_font = Font(bold=False, size=10)
    body_font = Font(size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    number_alignment = Alignment(horizontal="right", vertical="center")

    for row_index in range(1, data_start_row):
        for column_index in range(1, total_columns + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.border = thin_border
            cell.alignment = center
            if row_index == LCL_COST_HEADER_ROWS and column_index <= shipment_count:
                header_name = shipment_columns[column_index - 1]
                cell.font = (
                    header_font if header_name in BOLD_SHIPMENT_HEADERS else normal_header_font
                )
                cell.fill = shipment_fill
            else:
                cell.font = header_font if row_index < LCL_COST_HEADER_ROWS else normal_header_font
                cell.fill = cost_fill if column_index > shipment_count else shipment_fill

    for row_index in range(data_start_row, data_end_row + 1):
        for column_index in range(1, total_columns + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.border = thin_border
            cell.font = body_font
            if column_index <= shipment_count:
                cell.alignment = left
            elif _is_currency_column(column_index, shipment_count, cost_column_groups):
                cell.alignment = center
            else:
                cell.alignment = number_alignment
                if isinstance(cell.value, (int, float)):
                    cell.number_format = _number_format(cell.value)

    for column_index in range(1, total_columns + 1):
        column_letter = get_column_letter(column_index)
        max_length = 0
        for row_index in range(1, data_end_row + 1):
            value = worksheet.cell(row=row_index, column=column_index).value
            if value is None:
                continue
            max_length = max(max_length, len(str(value).split("\n")[0]))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 28)

    worksheet.freeze_panes = worksheet.cell(row=data_start_row, column=1)


def _write_rate_card_sheet(
    workbook: Workbook,
    rate_card_df: pd.DataFrame,
    column_groups: list[dict[str, object]],
    shipment_columns: list[str],
) -> None:
    worksheet = workbook.active
    worksheet.title = RATE_CARD_SHEET_NAME

    shipment_count = len(shipment_columns)
    cost_column_groups: list[tuple[int, ...]] = []
    next_column = shipment_count + 1

    for meta in column_groups:
        column_count = _column_group_width(meta)
        group_columns = tuple(range(next_column, next_column + column_count))
        cost_column_groups.append(group_columns)
        currency_col = group_columns[0]

        for row_index in (1, 2):
            worksheet.merge_cells(
                start_row=row_index,
                start_column=currency_col,
                end_row=row_index,
                end_column=group_columns[-1],
            )

        row3_flat_label = meta.get("row3_flat_label")
        if column_count == 3 and row3_flat_label:
            worksheet.merge_cells(
                start_row=3,
                start_column=currency_col,
                end_row=3,
                end_column=currency_col + 1,
            )
            worksheet.cell(row=3, column=currency_col, value=meta["rate_by"])
            worksheet.cell(row=3, column=group_columns[-1], value=row3_flat_label)
        else:
            worksheet.merge_cells(
                start_row=3,
                start_column=currency_col,
                end_row=3,
                end_column=group_columns[-1],
            )
            worksheet.cell(row=3, column=currency_col, value=meta["rate_by"])

        worksheet.cell(row=1, column=currency_col, value=meta["cost_name"])
        worksheet.cell(row=2, column=currency_col, value=meta["apply_if"])
        worksheet.cell(row=LCL_COST_HEADER_ROWS, column=currency_col, value=CURRENCY_COLUMN_HEADER)

        rate_types = meta.get("rate_types")
        if rate_types:
            for offset, rate_type in enumerate(rate_types, start=1):
                worksheet.cell(
                    row=LCL_COST_HEADER_ROWS,
                    column=currency_col + offset,
                    value=rate_type,
                )
        else:
            worksheet.cell(
                row=LCL_COST_HEADER_ROWS,
                column=currency_col + 1,
                value=meta["rate_type"],
            )

        next_column += column_count

    for column_index, column_name in enumerate(shipment_columns, start=1):
        worksheet.cell(row=LCL_COST_HEADER_ROWS, column=column_index, value=column_name)

    data_start_row = LCL_COST_HEADER_ROWS + 1
    data_rows = list(dataframe_to_rows(rate_card_df, index=False, header=False))
    for row_offset, row in enumerate(data_rows, start=data_start_row):
        for column_offset, value in enumerate(row, start=1):
            worksheet.cell(row=row_offset, column=column_offset, value=value)

    total_columns = shipment_count + sum(_column_group_width(meta) for meta in column_groups)
    data_end_row = data_start_row + len(data_rows) - 1
    _apply_rate_card_formatting(
        worksheet=worksheet,
        shipment_columns=shipment_columns,
        total_columns=total_columns,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
        cost_column_groups=cost_column_groups,
    )


def _write_conditional_rules_sheet(workbook: Workbook, conditional_df: pd.DataFrame) -> None:
    worksheet = workbook.create_sheet(title=CONDITIONAL_RULES_SHEET_NAME)
    for column_index, column_name in enumerate(CONDITIONAL_RULES_COLUMNS, start=1):
        worksheet.cell(row=1, column=column_index, value=column_name)

    data_rows = list(dataframe_to_rows(conditional_df, index=False, header=False))
    for row_offset, row in enumerate(data_rows, start=2):
        for column_offset, value in enumerate(row, start=1):
            worksheet.cell(row=row_offset, column=column_offset, value=value)

    data_end_row = max(1, 1 + len(data_rows))
    apply_table_sheet_formatting(
        worksheet,
        header_row=1,
        column_count=len(CONDITIONAL_RULES_COLUMNS),
        data_end_row=data_end_row,
        column_max_widths={
            1: 24,
            2: 20,
            3: 80,
        },
    )


def resolve_lcl_carrier_slug(
    shipper: str,
    individual_selections: list[SubfolderSelection],
) -> str:
    carrier_codes: list[str] = []
    carrier_keys: list[str] = []
    for selection in individual_selections:
        carrier_code = carrier_code_from_filename(
            selection.file_path.name,
            shipper=shipper,
            flow="FCL",
        )
        if carrier_code:
            carrier_codes.append(carrier_code)
            continue
        carrier_key = detect_carrier_key(
            selection.file_path.name,
            shipper=shipper,
            flow="FCL",
        )
        if carrier_key:
            carrier_keys.append(carrier_key)

    if carrier_codes:
        return slugify("_".join(sorted(set(carrier_codes))))
    if carrier_keys:
        return slugify("_".join(sorted(set(carrier_keys))))
    return "unknown"


def build_output_rate_card_path(
    flow: str,
    shipper: str,
    carrier_slug: str | None = None,
) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    shipper_slug = slugify(shipper)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if carrier_slug:
        return OUTPUT_DIR / f"{flow}_{shipper_slug}_{carrier_slug}_rate_card_{timestamp}.xlsx"
    return OUTPUT_DIR / f"{flow}_{shipper_slug}_rate_card_{timestamp}.xlsx"


def save_lcl_rate_card(
    shipper: str,
    individual_selections: list[SubfolderSelection],
    processing_path: Path | None = None,
    output_path: Path | None = None,
) -> tuple[Path, pd.DataFrame, pd.DataFrame]:
    if processing_path and processing_path.exists():
        parsed_rows, shipment_columns = load_lcl_parsed_rows_from_processing(
            processing_path,
            individual_selections,
        )
        if not parsed_rows:
            parsed_rows, shipment_columns = load_lcl_parsed_rows(individual_selections)
    else:
        parsed_rows, shipment_columns = load_lcl_parsed_rows(individual_selections)

    if not parsed_rows:
        raise ValueError("No LCL Rate/LCL_Rates rows found in selected individual rate files.")

    rate_card_df, column_groups, shipment_columns = build_lcl_rate_card_dataframe(
        parsed_rows,
        shipment_columns,
    )
    conditional_df = build_conditional_cures_dataframe(parsed_rows)

    carrier_slug = resolve_lcl_carrier_slug(shipper, individual_selections)
    if output_path is None:
        output_path = build_output_rate_card_path("LCL", shipper, carrier_slug)

    accessorial_df = build_lcl_accessorial_costs_dataframe(shipper, individual_selections)

    workbook = Workbook()
    _write_rate_card_sheet(workbook, rate_card_df, column_groups, shipment_columns)
    _write_conditional_rules_sheet(workbook, conditional_df)
    write_accessorial_costs_sheet(workbook, accessorial_df)
    workbook.save(output_path)

    return output_path, rate_card_df, conditional_df


def resolve_shipper(shipper: str | None = None) -> str:
    if shipper:
        return shipper
    context = load_processing_context()
    if context and context.get("shipper"):
        return context["shipper"]
    return "unknown"
