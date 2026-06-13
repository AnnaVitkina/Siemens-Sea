"""Build Innomotics LCL rate card from Rate Card_LCL-like tabs."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from config import OUTPUT_DIR, RATE_CARD_SHEET_NAME
from extractor import SubfolderSelection, slugify

SHIPMENT_COLUMNS = [
    "Lane ID",
    "Origin Country",
    "Origin CFS Code",
    "Origin CFS Name",
    "Destination Country",
    "Destination CFS Code",
    "Valid to",
    "Valid from",
]


@dataclass(frozen=True)
class CostGroup:
    key: str
    cost_name: str
    apply_if: str
    rate_by: str


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"", "nan", "none"} else text


def _normalize_number(value: object) -> object:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace(",", ".")
    try:
        number = float(text)
        if float(number).is_integer():
            return int(number)
        return number
    except ValueError:
        return value


def _format_date(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return _normalize_text(value)
    return parsed.strftime("%d.%m.%Y")


def _detect_header_row(raw_df: pd.DataFrame) -> int:
    for idx in range(min(30, len(raw_df))):
        row_values = {_normalize_text(value).lower() for value in raw_df.iloc[idx].tolist()}
        if "lane id" in row_values and ("origin country" in row_values or "origin cfs code" in row_values):
            return idx
    raise ValueError("Could not detect Innomotics LCL header row.")


def _prepare_lcl_dataframe(file_path: Path, tab_name: str) -> pd.DataFrame:
    raw = pd.read_excel(file_path, sheet_name=tab_name, header=None)
    header_row = _detect_header_row(raw)
    df = pd.read_excel(file_path, sheet_name=tab_name, header=header_row)
    df = df.rename(columns=lambda c: _normalize_text(c))
    if "Unnamed: 0" in df.columns:
        df = df.drop(columns=["Unnamed: 0"])
    df = df.dropna(axis=1, how="all").dropna(axis=0, how="all")
    return df


def _find_column(columns: list[str], *candidates: str) -> str | None:
    lower_map = {col.lower(): col for col in columns}
    for candidate in candidates:
        col = lower_map.get(candidate.lower())
        if col:
            return col
    return None


def _resolve_actual_rate_column(columns: list[str]) -> str | None:
    for col in columns:
        low = col.lower()
        if "actual rate" in low and "geopolitical" not in low:
            return col
    return None


def build_innomotics_rate_card_dataframe(selections: list[SubfolderSelection]) -> tuple[pd.DataFrame, list[CostGroup]]:
    rows_by_key: dict[tuple, dict[str, object]] = {}
    costs_by_row: dict[tuple, dict[str, tuple[object, object]]] = {}
    groups_by_key: dict[str, CostGroup] = {}

    for selection in selections:
        for tab in selection.tabs:
            if "lcl" not in tab.lower():
                continue
            df = _prepare_lcl_dataframe(selection.file_path, tab)
            columns = list(df.columns)

            lane_col = _find_column(columns, "Lane ID")
            origin_country_col = _find_column(columns, "Origin Country")
            origin_cfs_code_col = _find_column(columns, "Origin CFS Code")
            origin_cfs_name_col = _find_column(columns, "Origin CFS Name")
            dest_country_col = _find_column(columns, "Destination Country")
            dest_cfs_code_col = _find_column(columns, "Destination CFS Code")
            valid_from_col = _find_column(columns, "Valid from")
            valid_to_col = _find_column(columns, "Valid until", "Valid to")
            currency_col = _find_column(columns, "Currency LCL")
            cost_type_col = _find_column(columns, "Cost type", "Cost Type")
            actual_rate_col = _resolve_actual_rate_column(columns)
            fallback_lcl_rate_col = _find_column(
                columns,
                "LCL Rate\nCost actual rate W/M\nincluding Bunker and ETS charges",
                "LCL Rate",
            )
            basis_col = _find_column(columns, "Calculation Basis")

            required_shipment = [
                lane_col,
                origin_country_col,
                origin_cfs_code_col,
                origin_cfs_name_col,
                dest_country_col,
                dest_cfs_code_col,
                valid_to_col,
                valid_from_col,
            ]
            if any(col is None for col in required_shipment):
                continue

            for _, src in df.iterrows():
                shipment = {
                    "Lane ID": _normalize_text(src.get(lane_col)),
                    "Origin Country": _normalize_text(src.get(origin_country_col)),
                    "Origin CFS Code": _normalize_text(src.get(origin_cfs_code_col)),
                    "Origin CFS Name": _normalize_text(src.get(origin_cfs_name_col)),
                    "Destination Country": _normalize_text(src.get(dest_country_col)),
                    "Destination CFS Code": _normalize_text(src.get(dest_cfs_code_col)),
                    "Valid to": _format_date(src.get(valid_to_col)),
                    "Valid from": _format_date(src.get(valid_from_col)),
                }
                if not shipment["Lane ID"]:
                    continue
                key = tuple(shipment[col] for col in SHIPMENT_COLUMNS)
                rows_by_key.setdefault(key, shipment)
                costs_by_row.setdefault(key, {})

                currency = _normalize_text(src.get(currency_col)) if currency_col else ""
                rate_value = None
                if actual_rate_col:
                    rate_value = _normalize_number(src.get(actual_rate_col))
                if rate_value is None and fallback_lcl_rate_col:
                    rate_value = _normalize_number(src.get(fallback_lcl_rate_col))
                if rate_value is None:
                    continue

                if cost_type_col:
                    raw_cost_type = _normalize_text(src.get(cost_type_col))
                    if not raw_cost_type:
                        continue
                    cost_name = (
                        "Transport cost (LCL rate)"
                        if raw_cost_type.lower() == "lcl rate"
                        else raw_cost_type
                    )
                    group_key = f"COST|{cost_name.upper()}"
                else:
                    cost_name = "Transport cost"
                    group_key = "COST|TRANSPORT_COST"

                rate_by = _normalize_text(src.get(basis_col)) if basis_col else ""
                groups_by_key[group_key] = CostGroup(
                    key=group_key,
                    cost_name=cost_name,
                    apply_if="",
                    rate_by=rate_by,
                )
                costs_by_row[key][group_key] = (currency, rate_value)

    if not rows_by_key:
        raise ValueError("No Innomotics Rate_Card_LCL rows found in selected tabs.")

    sorted_keys = sorted(rows_by_key.keys())
    shipment_df = pd.DataFrame([rows_by_key[k] for k in sorted_keys], columns=SHIPMENT_COLUMNS)

    ordered_groups = sorted(groups_by_key.values(), key=lambda g: g.cost_name.lower())
    cost_blocks: list[pd.DataFrame] = []
    rendered_groups: list[CostGroup] = []
    for group in ordered_groups:
        currencies: list[object] = []
        rates: list[object] = []
        for key in sorted_keys:
            currency, value = costs_by_row.get(key, {}).get(group.key, ("", None))
            currencies.append(currency)
            rates.append(value)
        if not any(v is not None for v in rates):
            continue
        cost_blocks.append(
            pd.DataFrame(
                {
                    f"{group.key}__currency": currencies,
                    f"{group.key}__rate": rates,
                }
            )
        )
        rendered_groups.append(group)

    rate_card_df = pd.concat([shipment_df, *cost_blocks], axis=1)
    return rate_card_df, rendered_groups


def _write_innomotics_sheet(workbook: Workbook, df: pd.DataFrame, groups: list[CostGroup]) -> None:
    ws = workbook.active
    ws.title = RATE_CARD_SHEET_NAME
    shipment_count = len(SHIPMENT_COLUMNS)
    next_col = shipment_count + 1
    group_ranges: list[tuple[int, int]] = []

    for group in groups:
        start = next_col
        end = start + 1
        group_ranges.append((start, end))
        for row in (1, 2, 3):
            ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
        ws.cell(1, start, group.cost_name)
        ws.cell(2, start, group.apply_if)
        ws.cell(3, start, group.rate_by)
        ws.cell(4, start, "Currency")
        ws.cell(4, start + 1, "p/unit")
        next_col = end + 1

    for idx, col in enumerate(SHIPMENT_COLUMNS, start=1):
        ws.cell(4, idx, col)

    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=5):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row_idx, col_idx, value=value)

    total_cols = max(next_col - 1, shipment_count)
    thin = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    ship_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    cost_fill = PatternFill(fill_type="solid", fgColor="E2EFDA")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    header_font = Font(bold=True, size=10)
    body_font = Font(size=10)
    normal_header = Font(size=10)

    data_end_row = 4 + len(df)
    for r in range(1, 5):
        for c in range(1, total_cols + 1):
            cell = ws.cell(r, c)
            cell.border = thin
            cell.alignment = center
            cell.font = header_font if r < 4 else (header_font if c <= shipment_count else normal_header)
            cell.fill = ship_fill if c <= shipment_count else cost_fill

    currency_cols = {start for start, _ in group_ranges}
    for r in range(5, data_end_row + 1):
        for c in range(1, total_cols + 1):
            cell = ws.cell(r, c)
            cell.border = thin
            cell.font = body_font
            if c <= shipment_count:
                cell.alignment = left
            elif c in currency_cols:
                cell.alignment = center
            else:
                cell.alignment = right

    for c in range(1, total_cols + 1):
        letter = get_column_letter(c)
        max_len = 0
        for r in range(1, data_end_row + 1):
            value = ws.cell(r, c).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value).split("\n")[0]))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 45)

    ws.freeze_panes = "A5"


def build_output_rate_card_path(flow: str, shipper: str) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return OUTPUT_DIR / f"{slugify(flow)}_{slugify(shipper)}_rate_card_{timestamp}.xlsx"


def _resolve_innomotics_carrier_slug(selections: list[SubfolderSelection]) -> str:
    names = " ".join(selection.file_path.name.upper() for selection in selections)
    if "DACHSER" in names:
        return "Dachser"
    if "DHL" in names:
        return "DHL"
    if "KUEHNE" in names or " KN_" in names or "_KN_" in names or names.startswith("KN_"):
        return "Kuehne"
    return "unknown"


def save_innomotics_rate_card(
    shipper: str,
    selections: list[SubfolderSelection],
    output_path: Path | None = None,
) -> tuple[Path, pd.DataFrame]:
    rate_card_df, groups = build_innomotics_rate_card_dataframe(selections)
    if output_path is None:
        carrier_slug = _resolve_innomotics_carrier_slug(selections)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = OUTPUT_DIR / (
            f"{slugify('Innomotics')}_{slugify(shipper)}_{slugify(carrier_slug)}_rate_card_{timestamp}.xlsx"
        )
    workbook = Workbook()
    _write_innomotics_sheet(workbook, rate_card_df, groups)
    workbook.save(output_path)
    return output_path, rate_card_df

